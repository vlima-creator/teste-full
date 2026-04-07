import re
from io import BytesIO
from typing import Dict, Optional, List, Tuple
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Dashboard Fulfillment Estratégico",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# Premium UI (CSS) - Design Moderno e Fluido - CORRIGIDO
# ============================================================

st.markdown(
    """
<style>
/* ===== IMPORTS ===== */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ===== ROOT VARIABLES ===== */
:root {
    --primary: #2e7d32; /* Verde Militar */
    --primary-light: #4caf50;
    --primary-dark: #1b5e20;
    --secondary: #ec4899;
    --accent: #06b6d4;
    --success: #10b981;
    --warning: #f59e0b;
    --danger: #ef4444;
    --dark: #000000; /* Fundo Preto */
    --dark-light: #111111;
    --dark-lighter: #222222;
    --text-primary: #ffffff;
    --text-secondary: #a1a1aa;
    --text-muted: #71717a;
    --glass-bg: rgba(255, 255, 255, 0.03);
    --glass-border: rgba(255, 255, 255, 0.1);
    --liquid-glass-border: rgba(255, 255, 255, 0.15);
    --gradient-primary: linear-gradient(135deg, #2e7d32 0%, #1b5e20 100%);
    --hover-green: rgba(46, 125, 50, 0.2);
}

/* ===== BASE STYLES ===== */
html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}

.stApp {
    background-color: #000000 !important;
    background-image: none !important;
}

/* ===== MAIN CONTAINER ===== */
.block-container {
    padding: 2rem 3rem 3rem 3rem !important;
    max-width: 1600px !important;
}

/* ===== SIDEBAR PREMIUM ===== */
section[data-testid="stSidebar"] {
    background-color: #000000 !important;
    border-right: 1px solid var(--glass-border) !important;
}

section[data-testid="stSidebar"] > div:first-child {
    padding: 1.5rem 1rem !important;
}

/* ===== TYPOGRAPHY ===== */
h1 {
    font-size: 2.5rem !important;
    font-weight: 800 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    letter-spacing: -0.03em !important;
    margin-bottom: 0.5rem !important;
}

h2 {
    font-size: 1.5rem !important;
    font-weight: 700 !important;
    color: var(--text-primary) !important;
}

h3 {
    font-size: 1.25rem !important;
    font-weight: 600 !important;
    color: var(--text-primary) !important;
}

/* ===== CUSTOM CLASSES ===== */
.premium-header {
    text-align: center;
    padding: 1rem 0 2rem 0;
    margin-bottom: 1rem;
}

.premium-header h1 {
    font-size: 3rem !important;
    margin-bottom: 0.5rem !important;
}

.premium-subtitle {
    color: var(--text-muted) !important;
    font-size: 1rem !important;
    font-weight: 400 !important;
}

.context-bar {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 2rem;
    padding: 0.75rem 1.5rem;
    background: var(--glass-bg);
    border: 1px solid var(--liquid-glass-border);
    backdrop-filter: blur(10px);
    border-radius: 100px;
    margin: 0 auto 2rem auto;
    max-width: fit-content;
}

.context-item .icon-box {
    background: rgba(255, 255, 255, 0.1);
    border: 1px solid rgba(255, 255, 255, 0.2);
    width: 28px;
    height: 28px;
    border-radius: 6px;
}

.context-item {
    display: flex;
    align-items: center;
    gap: 0.5rem;
    font-size: 0.875rem;
    color: var(--text-secondary) !important;
}

.context-item .label {
    color: var(--text-muted) !important;
}

.context-item .value {
    color: var(--text-primary) !important;
    font-weight: 600;
}

/* ===== GLASS CARDS ===== */
.glass-card {
    background: var(--glass-bg) !important;
    border: 1px solid var(--liquid-glass-border) !important;
    backdrop-filter: blur(12px);
    border-radius: 16px !important;
    padding: 1.5rem !important;
    transition: all 0.3s ease;
}

.glass-card:hover {
    background: var(--hover-green) !important;
    border-color: var(--primary) !important;
    transform: translateY(-2px);
}

/* ===== STREAMLIT METRICS OVERRIDE ===== */
[data-testid="stMetric"] {
    background: var(--glass-bg) !important;
    border: 1px solid var(--liquid-glass-border) !important;
    backdrop-filter: blur(12px);
    border-radius: 16px !important;
    padding: 1.25rem !important;
    transition: all 0.3s ease;
}

[data-testid="stMetric"]:hover {
    background: var(--hover-green) !important;
    border-color: var(--primary) !important;
}

[data-testid="stMetricLabel"] {
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
    color: var(--text-muted) !important;
}

[data-testid="stMetricValue"] {
    font-size: 1.75rem !important;
    font-weight: 700 !important;
    color: var(--text-primary) !important;
}

/* ===== TABS PREMIUM ===== */
.stTabs [data-baseweb="tab-list"] {
    background: var(--glass-bg) !important;
    border-radius: 12px !important;
    padding: 0.5rem !important;
    gap: 0.5rem !important;
    border: 1px solid var(--liquid-glass-border) !important;
}

.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 8px !important;
    padding: 0.75rem 1.5rem !important;
    font-weight: 500 !important;
    color: var(--text-secondary) !important;
    border: none !important;
    transition: all 0.2s ease;
}

.stTabs [data-baseweb="tab"] p {
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
}

.stTabs [data-baseweb="tab"] p::before {
    content: '';
    display: inline-block;
    width: 20px;
    height: 20px;
    background-color: white;
    mask-size: contain;
    mask-repeat: no-repeat;
    mask-position: center;
    -webkit-mask-size: contain;
    -webkit-mask-repeat: no-repeat;
    -webkit-mask-position: center;
}

/* Mapeamento de ícones para as abas */
.stTabs [data-baseweb="tab"]:nth-child(1) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/layout-dashboard.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/layout-dashboard.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(2) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/trending-up.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/trending-up.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(3) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/target.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/target.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(4) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/package.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/package.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(5) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/clapperboard.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/clapperboard.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(6) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/truck.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/truck.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(7) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/save.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/save.svg'); }
.stTabs [data-baseweb="tab"]:nth-child(8) p::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/help-circle.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/help-circle.svg'); }

.stTabs [data-baseweb="tab"]:hover {
    background: var(--hover-green) !important;
    color: white !important;
}

.stTabs [aria-selected="true"] {
    background: var(--primary) !important;
    color: white !important;
    font-weight: 600 !important;
}

.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] {
    display: none !important;
}

/* ===== BUTTONS PREMIUM ===== */
.stButton > button {
    background: var(--primary) !important;
    color: white !important;
    border: 1px solid var(--liquid-glass-border) !important;
    border-radius: 10px !important;
    padding: 0.75rem 1.5rem !important;
    font-weight: 600 !important;
    transition: all 0.3s ease !important;
}

.stButton > button:hover {
    background: var(--primary-dark) !important;
    border-color: var(--primary-light) !important;
    box-shadow: 0 0 15px var(--hover-green) !important;
}

.stDownloadButton > button {
    background: var(--dark-lighter) !important;
    color: white !important;
    border: 1px solid var(--liquid-glass-border) !important;
    border-radius: 10px !important;
    padding: 0.75rem 1.5rem !important;
    font-weight: 600 !important;
    transition: all 0.3s ease !important;
}

.stDownloadButton > button:hover {
    background: var(--hover-green) !important;
    border-color: var(--primary) !important;
}

/* ===== INPUTS PREMIUM ===== */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stSelectbox > div > div > div {
    background: var(--dark-light) !important;
    border: 1px solid var(--liquid-glass-border) !important;
    border-radius: 10px !important;
    color: var(--text-primary) !important;
    transition: all 0.3s ease !important;
}

.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus {
    border-color: var(--primary) !important;
    box-shadow: 0 0 0 2px var(--hover-green) !important;
}

/* ===== FILE UPLOADER PREMIUM ===== */
[data-testid="stFileUploader"] {
    background: var(--glass-bg) !important;
    border: 2px dashed var(--liquid-glass-border) !important;
    border-radius: 12px !important;
    padding: 1rem !important;
    transition: all 0.3s ease;
}

[data-testid="stFileUploader"]:hover {
    border-color: var(--primary) !important;
    background: var(--hover-green) !important;
}

/* ===== EXPANDER PREMIUM ===== */
.streamlit-expanderHeader {
    background: var(--glass-bg) !important;
    border: 1px solid var(--liquid-glass-border) !important;
    border-radius: 12px !important;
    color: var(--text-primary) !important;
    font-weight: 600 !important;
    transition: all 0.3s ease !important;
}

.streamlit-expanderHeader:hover {
    background: var(--hover-green) !important;
    border-color: var(--primary) !important;
}

/* ===== DIVIDER ===== */
.hr {
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.1), transparent);
    margin: 1.5rem 0;
}

/* ===== INFO BOX ===== */
.info-box {
    background: rgba(46, 125, 50, 0.05);
    border: 1px solid var(--liquid-glass-border);
    border-radius: 12px;
    padding: 1rem 1.25rem;
    margin: 1rem 0;
}

.info-box p {
    color: var(--text-secondary) !important;
    margin: 0;
}

/* ===== SCROLLBAR PREMIUM ===== */
::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}

::-webkit-scrollbar-track {
    background: var(--dark);
    border-radius: 4px;
}

::-webkit-scrollbar-thumb {
    background: var(--dark-lighter);
    border-radius: 4px;
}

/* ===== LOGO AREA ===== */
.logo-area {
    text-align: center;
    padding: 1rem 0 1.5rem 0;
    border-bottom: 1px solid var(--glass-border);
    margin-bottom: 1.5rem;
}

.logo-text {
    font-size: 1.5rem;
    font-weight: 800;
    color: #ffffff;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
}

.logo-text::before {
    content: '';
    display: inline-block;
    width: 24px;
    height: 24px;
    background-color: white;
    -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/bar-chart-3.svg');
    mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/bar-chart-3.svg');
    -webkit-mask-size: contain;
    mask-size: contain;
    -webkit-mask-repeat: no-repeat;
    mask-repeat: no-repeat;
}

.logo-subtitle {
    font-size: 0.75rem;
    color: var(--text-muted) !important;
    text-transform: uppercase;
    letter-spacing: 0.15em;
    margin-top: 0.25rem;
}

/* ===== SIDEBAR SECTIONS ===== */
.sidebar-section-title {
    font-size: 0.7rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: var(--text-muted) !important;
    margin-bottom: 0.75rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

.sidebar-section-title::before {
    content: '';
    display: inline-block;
    width: 14px;
    height: 14px;
    background-color: var(--text-muted);
    mask-size: contain;
    mask-repeat: no-repeat;
    mask-position: center;
    -webkit-mask-size: contain;
    -webkit-mask-repeat: no-repeat;
    -webkit-mask-position: center;
}

/* Mapeamento de ícones para a sidebar */
div:has(> .sidebar-section-title:contains("Arquivos")) .sidebar-section-title::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/folder.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/folder.svg'); }
div:has(> .sidebar-section-title:contains("Movimentações")) .sidebar-section-title::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/package.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/package.svg'); }
div:has(> .sidebar-section-title:contains("Vendas")) .sidebar-section-title::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/dollar-sign.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/dollar-sign.svg'); }
div:has(> .sidebar-section-title:contains("Parâmetros")) .sidebar-section-title::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/settings.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/settings.svg'); }
div:has(> .sidebar-section-title:contains("Oportunidades")) .sidebar-section-title::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/target.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/target.svg'); }
div:has(> .sidebar-section-title:contains("Simulação")) .sidebar-section-title::before { -webkit-mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/flask-conical.svg'); mask-image: url('https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/flask-conical.svg'); }

/* ===== HEALTH INDICATORS ===== */
.health-card {
    background: var(--glass-bg);
    border: 1px solid var(--liquid-glass-border);
    backdrop-filter: blur(12px);
    border-radius: 16px;
    padding: 1.25rem;
    position: relative;
    overflow: hidden;
    transition: all 0.3s ease;
}

.health-card:hover {
    background: var(--hover-green);
    border-color: var(--primary);
}

.health-card::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 4px;
}

.health-card.critical::before {
    background: #ef4444;
}

.health-card.warning::before {
    background: #f59e0b;
}

.health-card.good::before {
    background: #10b981;
}

.health-card.info::before {
    background: #2e7d32;
}

.health-label {
    font-size: 0.7rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: var(--text-muted);
    margin-bottom: 0.5rem;
}

.health-value {
    font-size: 2rem;
    font-weight: 700;
    color: var(--text-primary);
    line-height: 1.2;
}

.health-subtitle {
    font-size: 0.8rem;
    color: var(--text-secondary);
    margin-top: 0.25rem;
}

/* ===== ALERT BADGES ===== */
.alert-badge {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    padding: 0.5rem 1rem;
    border-radius: 100px;
    font-size: 0.8rem;
    font-weight: 600;
    border: 1px solid var(--liquid-glass-border);
    backdrop-filter: blur(8px);
}

.alert-badge.critical {
    background: rgba(239, 68, 68, 0.1);
    color: #ef4444;
}

.alert-badge.warning {
    background: rgba(245, 158, 11, 0.1);
    color: #f59e0b;
}

.alert-badge.success {
    background: rgba(16, 185, 129, 0.1);
    color: #10b981;
}

.alert-badge.info {
    background: rgba(46, 125, 50, 0.1);
    color: #4caf50;
}

/* ===== PRIORITY LIST ===== */
.priority-item {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1rem;
    background: var(--glass-bg);
    border: 1px solid var(--liquid-glass-border);
    border-radius: 12px;
    margin-bottom: 0.75rem;
    transition: all 0.3s ease;
}

.priority-item:hover {
    background: var(--hover-green);
    border-color: var(--primary);
}

.priority-rank {
    width: 32px;
    height: 32px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: var(--primary);
    border-radius: 8px;
    font-weight: 700;
    font-size: 0.875rem;
    color: white;
}

.priority-info {
    flex: 1;
    margin-left: 1rem;
}

.priority-sku {
    font-weight: 600;
    color: var(--text-primary);
}

.priority-detail {
    font-size: 0.75rem;
    color: var(--text-muted);
}

.priority-action {
    text-align: right;
}

.priority-qty {
    font-size: 1.25rem;
    font-weight: 700;
    color: var(--primary-light);
}

.priority-label {
    font-size: 0.7rem;
    color: var(--text-muted);
}

/* ===== CHART CONTAINER ===== */
.chart-container {
    background: var(--glass-bg);
    border: 1px solid var(--liquid-glass-border);
    backdrop-filter: blur(12px);
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    transition: all 0.3s ease;
}

.chart-container:hover {
    border-color: var(--primary);
}

.chart-title {
    font-size: 1rem;
    font-weight: 600;
    color: var(--text-primary);
    margin-bottom: 1rem;
}

/* ===== SUMMARY GRID ===== */
.summary-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 1rem;
    margin-bottom: 1.5rem;
}

/* ===== PROGRESS BAR ===== */
.progress-container {
    background: var(--dark-lighter);
    border-radius: 100px;
    height: 8px;
    overflow: hidden;
    margin-top: 0.5rem;
}

.progress-bar {
    height: 100%;
    border-radius: 100px;
    transition: width 0.3s ease;
}

.progress-bar.critical {
    background: #ef4444;
}

.progress-bar.warning {
    background: #f59e0b;
}

.progress-bar.good {
    background: #10b981;
}

/* ===== ABC BADGES ===== */
.abc-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 28px;
    height: 28px;
    border-radius: 6px;
    font-weight: 700;
    font-size: 0.8rem;
}

.abc-badge.a {
    background: #10b981;
    color: white;
}

.abc-badge.b {
    background: #f59e0b;
    color: white;
}

.abc-badge.c {
    background: #71717a;
    color: white;
}

/* ===== ICON STANDARDIZATION ===== */
.icon-white {
    filter: brightness(0) invert(1);
    width: 20px;
    height: 20px;
    display: inline-block;
    vertical-align: middle;
}

.icon-box {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 32px;
    height: 32px;
    border: 1px solid var(--liquid-glass-border);
    border-radius: 8px;
    margin-right: 8px;
    background: rgba(255, 255, 255, 0.05);
}

/* ===== RESPONSIVE ===== */
@media (max-width: 768px) {
    .block-container {
        padding: 1rem !important;
    }
    
    h1 {
        font-size: 1.75rem !important;
    }
    
    .context-bar {
        flex-direction: column;
        gap: 0.5rem;
        padding: 1rem;
    }
}
</style>
""",
    unsafe_allow_html=True,
)

# ============================================================
# Helpers
# ============================================================

def _clean_str(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    return str(x).strip()


def _to_number(x):
    if x is None:
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none", "-"}:
        return np.nan

    s = re.sub(r"[Rr]\$|\s", "", s)
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan


def _format_int(n) -> str:
    try:
        return f"{int(round(float(n))):,}".replace(",", ".")
    except Exception:
        return str(n)


def _format_brl(x) -> str:
    try:
        v = float(x)
        s = f"{v:,.2f}"
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {s}"
    except Exception:
        return str(x)


def _extract_first_int(text: str) -> Optional[int]:
    m = re.search(r"(\d+)", text or "")
    return int(m.group(1)) if m else None


def _find_row_index(df: pd.DataFrame, needle: str) -> Optional[int]:
    needle = needle.strip()
    for i in range(len(df)):
        row = df.iloc[i].astype(str)
        if (row == needle).any():
            return i
    return None


def _safe_get(df: pd.DataFrame, r: int, c: int) -> str:
    try:
        return _clean_str(df.iloc[r, c])
    except Exception:
        return ""


def _normalize_colname(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _detect_column(columns: List[str], patterns: List[str]) -> Optional[str]:
    cols = list(columns)
    norm = {c: _normalize_colname(c) for c in cols}
    for p in patterns:
        p2 = p.lower()
        for c in cols:
            if p2 in norm[c]:
                return c
    return None


def _make_unique_columns(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, c in enumerate(cols):
        c0 = _clean_str(c)
        if c0 == "":
            c0 = f"col_{i}"
        if c0 in seen:
            seen[c0] += 1
            out.append(f"{c0}_{seen[c0]}")
        else:
            seen[c0] = 1
            out.append(c0)
    return out


def _unnamed_ratio(cols: List[str]) -> float:
    if not cols:
        return 1.0
    cnt = 0
    for c in cols:
        s = str(c)
        if s.startswith("Unnamed") or s.strip() == "":
            cnt += 1
    return cnt / max(1, len(cols))


# ============================================================
# SAFE RENDER (evita crash PyArrow) - CORRIGIDO
# ============================================================
def safe_df(df: pd.DataFrame) -> pd.DataFrame:
    """Converte DataFrame para formato seguro para exibição."""
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    # Limpa nomes de colunas e garante unicidade
    cols = []
    for i, c in enumerate(list(out.columns)):
        s = "" if c is None else str(c)
        s = s.replace("\n", " ").replace("\r", " ").strip()
        if s == "":
            s = f"col_{i}"
        cols.append(s)
    # Garante que todas as colunas sejam únicas
    cols = _make_unique_columns(cols)
    out.columns = cols

    out = out.reset_index(drop=True)

    # Converte células para tipos seguros
    for c in out.columns:
        try:
            # Tenta converter para numérico primeiro
            numeric_col = pd.to_numeric(out[c], errors='coerce')
            if numeric_col.notna().sum() > 0.5 * len(numeric_col):
                out[c] = numeric_col
                continue
        except:
            pass
        
        # Converte para string se não for numérico
        try:
            out[c] = out[c].apply(lambda x: "" if pd.isna(x) else str(x))
        except (ValueError, TypeError):
            # Fallback: converte diretamente para string
            out[c] = out[c].astype(str)

    return out


def to_xlsx_bytes(dfs_dict: Dict[str, pd.DataFrame]) -> bytes:
    """Gera um arquivo XLSX formatado a partir de um dicionário de DataFrames."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs_dict.items():
            # Limitar nome da aba a 31 caracteres
            safe_sheet_name = str(sheet_name)[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            ws = writer.sheets[safe_sheet_name]

            # Estilização do cabeçalho (Verde Militar do app)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Congelar painel superior
            ws.freeze_panes = "A2"

            # Ajuste automático de largura das colunas
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_length = 0

                # Verificar cabeçalho
                header_val = str(col_cells[0].value) if col_cells[0].value else ""
                max_length = len(header_val)

                # Verificar amostra de dados (primeiras 100 linhas) para performance
                for cell in col_cells[1:101]:
                    try:
                        if cell.value:
                            val_len = len(str(cell.value))
                            if val_len > max_length:
                                max_length = val_len
                    except:
                        pass

                # Definir largura com margem, limites entre 10 e 50
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)
    return output.getvalue()


def show_df(df: pd.DataFrame, use_container_width: bool = True, height: int = 400):
    """Exibe DataFrame de forma segura."""
    if df is None or df.empty:
        st.info("Nenhum dado disponível.")
        return
    
    d = safe_df(df)
    
    try:
        st.dataframe(d, use_container_width=use_container_width, height=height)
    except Exception as e:
        # Fallback: converte tudo para string
        d_str = d.astype(str)
        try:
            st.dataframe(d_str, use_container_width=use_container_width, height=height)
        except:
            st.error(f"Erro ao exibir tabela: {e}")


def show_df_minimal(df: pd.DataFrame, height: int = 400):
    """Render minimalista sem column_config para evitar erros."""
    if df is None or df.empty:
        st.info("Nenhum dado disponível.")
        return
    
    d = safe_df(df)
    
    try:
        st.dataframe(d, use_container_width=True, height=height)
    except Exception as e:
        # Fallback: converte tudo para string
        d_str = d.astype(str)
        try:
            st.dataframe(d_str, use_container_width=True, height=height)
        except:
            st.error(f"Erro ao exibir tabela: {e}")


# ============================================================
# Sazonalidade Brasil
# ============================================================

def _nth_weekday_of_month(year: int, month: int, weekday: int, n: int) -> date:
    d = date(year, month, 1)
    shift = (weekday - d.weekday()) % 7
    d = d + timedelta(days=shift)
    d = d + timedelta(weeks=n - 1)
    return d


def _last_weekday_of_month(year: int, month: int, weekday: int) -> date:
    if month == 12:
        d = date(year, 12, 31)
    else:
        d = date(year, month + 1, 1) - timedelta(days=1)
    shift = (d.weekday() - weekday) % 7
    return d - timedelta(days=shift)


def brazil_retail_events(year: int) -> List[Tuple[str, date]]:
    """Calendário Comercial Completo de Marketplaces 2026 - Atualizado"""
    
    # Calcular Páscoa (varia a cada ano)
    # Algoritmo de Meeus/Jones/Butcher para cálculo da Páscoa
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    pascoa = date(year, month, day)
    carnaval = pascoa - timedelta(days=47)  # Carnaval é 47 dias antes da Páscoa
    
    return [
        # Janeiro
        ("Ano Novo / Liquidações", date(year, 1, 1)),
        ("Dia do Leitor", date(year, 1, 7)),
        ("Dia da Gula", date(year, 1, 26)),
        
        # Fevereiro
        ("Datas Duplas 2.2", date(year, 2, 2)),
        ("Volta às Aulas", date(year, 2, 4)),
        ("Carnaval", carnaval),
        ("Valentine's Day", date(year, 2, 14)),
        ("Dia do Esportista", date(year, 2, 19)),
        
        # Março
        ("Dia da Mulher", date(year, 3, 8)),
        ("Dia do Consumidor", date(year, 3, 15)),
        ("Início do Outono", date(year, 3, 20)),
        
        # Abril
        ("Páscoa", pascoa),
        ("Dia Mundial do Café", date(year, 4, 14)),
        ("Dia Mundial do Livro", date(year, 4, 23)),
        ("Dia do Frete Grátis", date(year, 4, 28)),
        
        # Maio
        ("Dia do Trabalho", date(year, 5, 1)),
        ("Dia das Mães", _nth_weekday_of_month(year, 5, weekday=6, n=2)),
        ("Dia do Orgulho Nerd", date(year, 5, 25)),
        
        # Junho
        ("Datas Duplas 6.6", date(year, 6, 6)),
        ("Dia dos Namorados", date(year, 6, 12)),
        ("Início do Inverno", date(year, 6, 21)),
        ("São João", date(year, 6, 24)),
        ("Dia do Orgulho LGBTQIAP+", date(year, 6, 28)),
        
        # Julho - PICO MÁXIMO
        ("Datas Duplas 7.7 (Aniversário Shopee)", date(year, 7, 7)),
        ("Dia Mundial do Chocolate", date(year, 7, 7)),
        ("Dia Mundial do Rock", date(year, 7, 13)),
        ("Dia dos Amigos", date(year, 7, 20)),
        ("Dia dos Avós", date(year, 7, 26)),
        
        # Agosto
        ("Datas Duplas 8.8", date(year, 8, 8)),
        ("Dia dos Pais", _nth_weekday_of_month(year, 8, weekday=6, n=2)),
        ("Dia do Estudante", date(year, 8, 11)),
        ("Dia dos Solteiros", date(year, 8, 15)),
        
        # Setembro
        ("Semana do Brasil", date(year, 9, 1)),
        ("Datas Duplas 9.9", date(year, 9, 9)),
        ("Dia do Cliente", date(year, 9, 15)),
        ("Início da Primavera", date(year, 9, 23)),
        
        # Outubro
        ("Dia Mundial dos Animais", date(year, 10, 4)),
        ("Datas Duplas 10.10", date(year, 10, 10)),
        ("Dia das Crianças", date(year, 10, 12)),
        ("Halloween", date(year, 10, 31)),
        
        # Novembro - PICOS MÁXIMOS
        ("Dia do Veganismo", date(year, 11, 1)),
        ("Datas Duplas 11.11 (Single's Day)", date(year, 11, 11)),
        ("Dia da Consciência Negra", date(year, 11, 20)),
        ("Black Friday", _last_weekday_of_month(year, 11, weekday=4)),
        ("Cyber Monday", _last_weekday_of_month(year, 11, weekday=4) + timedelta(days=3)),
        
        # Dezembro
        ("Datas Duplas 12.12 (Liquida de Natal)", date(year, 12, 12)),
        ("Natal", date(year, 12, 25)),
        ("Réveillon", date(year, 12, 31)),
    ]


# Eventos de PICO MÁXIMO (preparar estoque triplicado)
PICOS_MAXIMOS = [
    "Datas Duplas 11.11 (Single's Day)",
    "Black Friday",
    "Datas Duplas 7.7 (Aniversário Shopee)",
    "Dia das Mães",
]

# Eventos de PICO ALTO (preparar estoque duplicado)
PICOS_ALTOS = [
    "Dia das Crianças",
    "Datas Duplas 10.10",
    "Dia dos Namorados",
    "Natal",
    "Datas Duplas 9.9",
]


# ============================================================
# Detecção de Categoria Vestuário/Moda e Sazonalidade por Estação
# ============================================================

# Palavras-chave para detectar categoria de vestuário/moda
FASHION_KEYWORDS = [
    # Roupas gerais
    "camiseta", "camisa", "blusa", "regata", "top", "cropped", "body",
    "calça", "jeans", "legging", "shorts", "bermuda", "saia", "vestido",
    "casaco", "jaqueta", "blazer", "moletom", "cardigan", "colete", "parka",
    "suéter", "tricot", "malha", "lã", "fleece",
    # Calçados
    "sapato", "tênis", "sandália", "chinelo", "bota", "sapatilha", "mocassim",
    "rasteira", "tamanco", "mule", "slip on", "sneaker", "plataforma",
    # Acessórios
    "bolsa", "mochila", "carteira", "cinto", "lenço", "cachecol", "luva",
    "chapéu", "boné", "gorro", "óculos", "relógio", "pulseira", "colar",
    "brinco", "anel", "bijuteria", "jóia",
    # Moda íntima e praia
    "lingerie", "sutiã", "calcinha", "cueca", "boxer", "pijama", "robe",
    "biquíni", "maiô", "saida de praia", "canga", "sunga",
    # Categorias gerais
    "moda", "roupa", "vestuário", "fashion", "look", "outfit", "coleção",
    "feminino", "masculino", "infantil", "plus size", "fitness",
]

# Sazonalidade de Vestuário por Estação
FASHION_SEASONS = {
    "verao": {
        "nome": "Verão",
        "inicio": (12, 21),
        "fim": (3, 20),
        "emoji": "☀️",
        "produtos_destaque": [
            "Roupas leves e frescas",
            "Chinelos e sandálias",
            "Óculos de sol",
            "Biquínis e maiôs",
            "Shorts e bermudas",
            "Regatas e cropped",
            "Vestidos fluidos",
            "Protetor solar",
        ],
        "recomendacao_estoque": "Priorize peças leves, cores claras e estampas tropicais. Alta demanda por moda praia.",
        "cor_tema": "#FFD700",
    },
    "outono": {
        "nome": "Outono",
        "inicio": (3, 20),
        "fim": (6, 21),
        "emoji": "🍂",
        "produtos_destaque": [
            "Meia estação e transição",
            "Casacos leves e cardigans",
            "Sapatos fechados",
            "Jaquetas jeans",
            "Calças e leggings",
            "Blusas manga longa",
            "Cores terrosas e neutras",
            "Camadas e sobreposições",
        ],
        "recomendacao_estoque": "Transição de coleção. Liquide verão e prepare peças de meia estação.",
        "cor_tema": "#D2691E",
    },
    "inverno": {
        "nome": "Inverno",
        "inicio": (6, 21),
        "fim": (9, 22),
        "emoji": "❄️",
        "produtos_destaque": [
            "Casacos e jaquetas pesadas",
            "Botas e coturnos",
            "Roupas quentes e térmicas",
            "Acessórios (lenços, luvas, gorros)",
            "Moletom e fleece",
            "Tricot e malhas",
            "Cores escuras e sóbrias",
            "Pijamas e robes",
        ],
        "recomendacao_estoque": "Pico de demanda por agasalhos. Estoque 2x em casacos e botas.",
        "cor_tema": "#4169E1",
    },
    "primavera": {
        "nome": "Primavera",
        "inicio": (9, 22),
        "fim": (12, 21),
        "emoji": "🌸",
        "produtos_destaque": [
            "Flores e cores vibrantes",
            "Lançamentos de coleção",
            "Peças leves e fluidas",
            "Estampas florais",
            "Vestidos e saias",
            "Sandálias e rasteiras",
            "Tons pastéis",
            "Tecidos naturais",
        ],
        "recomendacao_estoque": "Lançamento de coleção verão. Aposte em novidades e estampas.",
        "cor_tema": "#FF69B4",
    },
}


def detect_fashion_category(df: pd.DataFrame) -> Dict:
    """
    Detecta se o relatório contém produtos de vestuário/moda
    baseado nas descrições dos produtos.
    """
    if df is None or df.empty:
        return {"is_fashion": False, "confidence": 0, "fashion_products": 0, "total_products": 0}
    
    # Procurar coluna de produto/descrição
    product_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if any(kw in col_lower for kw in ["produto", "descri", "titulo", "nome", "item"]):
            product_col = col
            break
    
    if product_col is None:
        return {"is_fashion": False, "confidence": 0, "fashion_products": 0, "total_products": 0}
    
    # Contar produtos de moda
    total_products = len(df)
    fashion_count = 0
    fashion_items = []
    
    for idx, row in df.iterrows():
        product_text = str(row[product_col]).lower()
        for keyword in FASHION_KEYWORDS:
            if keyword.lower() in product_text:
                fashion_count += 1
                fashion_items.append(row[product_col])
                break
    
    # Calcular confiança (% de produtos de moda)
    confidence = (fashion_count / total_products * 100) if total_products > 0 else 0
    
    # Considera categoria de moda se mais de 30% dos produtos são de vestuário
    is_fashion = confidence >= 30
    
    return {
        "is_fashion": is_fashion,
        "confidence": round(confidence, 1),
        "fashion_products": fashion_count,
        "total_products": total_products,
        "sample_items": fashion_items[:10],  # Amostra de 10 itens
    }


def get_current_fashion_season(today: date) -> Dict:
    """
    Retorna a estação atual do ano para moda/vestuário.
    """
    month = today.month
    day = today.day
    
    # Determinar estação baseado na data
    if (month == 12 and day >= 21) or (month in [1, 2]) or (month == 3 and day < 20):
        season_key = "verao"
    elif (month == 3 and day >= 20) or (month in [4, 5]) or (month == 6 and day < 21):
        season_key = "outono"
    elif (month == 6 and day >= 21) or (month in [7, 8]) or (month == 9 and day < 22):
        season_key = "inverno"
    else:  # (month == 9 and day >= 22) or (month in [10, 11]) or (month == 12 and day < 21)
        season_key = "primavera"
    
    season = FASHION_SEASONS[season_key].copy()
    season["key"] = season_key
    
    # Calcular dias restantes na estação
    fim_mes, fim_dia = season["fim"]
    if season_key == "verao" and today.month in [1, 2, 3]:
        fim_date = date(today.year, fim_mes, fim_dia)
    elif season_key == "verao":
        fim_date = date(today.year + 1, fim_mes, fim_dia)
    else:
        fim_date = date(today.year, fim_mes, fim_dia)
        if fim_date < today:
            fim_date = date(today.year + 1, fim_mes, fim_dia)
    
    season["dias_restantes"] = (fim_date - today).days
    
    # Determinar próxima estação
    next_seasons = {"verao": "outono", "outono": "inverno", "inverno": "primavera", "primavera": "verao"}
    season["proxima_estacao"] = FASHION_SEASONS[next_seasons[season_key]]["nome"]
    
    return season


def seasonal_calibration(today: date) -> Dict:
    """Calibração sazonal baseada no calendário comercial completo."""
    
    # Buscar todos os eventos futuros
    candidates = []
    for y in (today.year, today.year + 1):
        for name, d in brazil_retail_events(y):
            if d >= today:
                candidates.append((name, d))
    candidates.sort(key=lambda x: x[1])
    
    # Sempre ter um próximo evento
    if not candidates:
        # Se não houver eventos futuros no ano atual, buscar no próximo
        for name, d in brazil_retail_events(today.year + 1):
            candidates.append((name, d))
        candidates.sort(key=lambda x: x[1])
    
    next_name = candidates[0][0] if candidates else "Sem eventos próximos"
    next_date = candidates[0][1] if candidates else today + timedelta(days=999)
    days_to = (next_date - today).days
    
    # Determinar modo e cobertura baseado no tipo de evento e proximidade
    if next_name in PICOS_MAXIMOS:
        if days_to <= 60:
            return {
                "mode": "PICO MÁXIMO - Estoque Triplicado",
                "cover_days": 75,
                "context": f"Pré-{next_name}! Prepare estoque 3x. Maior pico do período.",
                "next_event": next_name,
                "days_to_event": days_to,
                "priority": "CRITICO",
            }
    
    if next_name in PICOS_ALTOS:
        if days_to <= 45:
            return {
                "mode": "PICO ALTO - Estoque Duplicado",
                "cover_days": 60,
                "context": f"Pré-{next_name}! Prepare estoque 2x. Alta demanda esperada.",
                "next_event": next_name,
                "days_to_event": days_to,
                "priority": "ALTO",
            }
    
    # Datas Duplas (Spike Days)
    if "Datas Duplas" in next_name:
        if days_to <= 30:
            return {
                "mode": "Spike Day - Estoque Aumentado",
                "cover_days": 50,
                "context": f"{next_name} em {days_to} dias. Prepare estoque 1.5x e cupons.",
                "next_event": next_name,
                "days_to_event": days_to,
                "priority": "MEDIO",
            }
    
    # Período pós-sazonal (Janeiro/Fevereiro - foco em liquidar)
    if today.month in (1, 2) and today.day <= 15:
        return {
            "mode": "Pós-sazonal - Liquidação",
            "cover_days": 35,
            "context": "Período de liquidação. Foco em giro rápido e evitar estoque parado.",
            "next_event": next_name,
            "days_to_event": days_to,
            "priority": "BAIXO",
        }
    
    # Qualquer evento próximo (até 45 dias)
    if days_to <= 45:
        return {
            "mode": "Pré-data comercial",
            "cover_days": 50,
            "context": f"Próximo evento: {next_name} em {days_to} dias.",
            "next_event": next_name,
            "days_to_event": days_to,
            "priority": "NORMAL",
        }
    
    # Cenário padrão
    return {
        "mode": "Cenário padrão",
        "cover_days": 45,
        "context": "Operação padrão, equilibrando giro e cobertura.",
        "next_event": next_name,
        "days_to_event": days_to,
        "priority": "NORMAL",
    }


# ============================================================
# UI: Componentes
# ============================================================

def show_guia_uso():
    st.markdown("""
    <style>
    /* Estilização para o Acordeão (st.expander) conforme o print */
    .stExpander {
        background: #000000 !important;
        border: 1px solid #222222 !important;
        border-radius: 8px !important;
        margin-bottom: 10px !important;
    }
    .stExpander > details > summary {
        background: #000000 !important;
        color: #ffffff !important;
        padding: 15px 20px !important;
        font-weight: 500 !important;
        border-radius: 8px !important;
    }
    .stExpander > details > summary:hover {
        background: #111111 !important;
    }
    .stExpander > details[open] > summary {
        border-bottom: 1px solid #222222 !important;
        border-radius: 8px 8px 0 0 !important;
    }
    .stExpander > details > div[role="region"] {
        background: #000000 !important;
        color: #a1a1aa !important;
        padding: 20px !important;
        border-radius: 0 0 8px 8px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='glass-card' style='margin-bottom: 2rem;'>
        <h2 style='margin-bottom: 0.5rem;'>📘 Guia de Uso - Dashboard Fulfillment Estratégico</h2>
        <p style='color: var(--text-secondary);'>Este guia explica como utilizar o dashboard, quais relatórios são necessários, como configurar os parâmetros e como os cálculos são realizados.</p>
    </div>
    """, unsafe_allow_html=True)

    # 1. Como Começar - Passo a Passo
    with st.expander("Como Começar - Passo a Passo", expanded=False):
        st.markdown("""
        1. **Upload de Arquivos:** No menu lateral, faça o upload dos relatórios do Mercado Livre.
           - **Relatório Geral de Estoque (Obrigatório):** Status atual do estoque no Full.
           - **Consolidado de Movimentações (Opcional):** Auditoria de entradas e saídas.
           - **Relatório de Vendas 30d (Opcional):** Velocidade de vendas e sugestões de reposição.
        2. **Configurações:** Ajuste os parâmetros na sidebar para personalizar a análise.
        3. **Navegação:** Utilize as abas superiores para alternar entre as visões estratégicas.
        """)

    # 2. Localizando os Relatórios
    with st.expander("Localizando os Relatórios", expanded=False):
        st.markdown("""
        Baixe os arquivos `.xlsx` diretamente do seu painel do Mercado Livre:
        
        | Relatório | Caminho no Mercado Livre |
        | :--- | :--- |
        | **Geral de Estoque** | Logística > Full > Gestão de estoque > Exportar Excel |
        | **Movimentações** | Logística > Full > Gestão de estoque > Relatório Consolidado > Exportar |
        | **Vendas (30 dias)** | Vendas > Anúncios > Relatórios de vendas > 30 dias > Exportar |
        """)

    # 3. O que você encontra em cada Aba
    with st.expander("O que você encontra em cada Aba", expanded=False):
        st.markdown("""
        #### 📈 Análise Estratégica
        Visão executiva da saúde do seu estoque. Aqui você identifica:
        - **Taxa de Ruptura:** SKUs que estão vendendo mas não têm estoque.
        - **Peso Morto:** Itens parados no estoque sem vendas recentes.
        - **Cobertura Média:** Quantos dias, em média, seu estoque atual suporta.
        - **Gráficos de Tendência:** Distribuição de estoque por categoria ou status.

        #### 🎯 Oportunidades
        Identificação de potencial de crescimento:
        - **Produtos fora do Full:** SKUs que vendem bem no seu estoque próprio mas ainda não foram enviados para o Full.
        - **Giro sem Estoque:** Itens que esgotaram recentemente mas possuem alto histórico de vendas.

        #### 📋 Painel de Ações
        A central de comando para o dia a dia:
        - **Ações Recomendadas:** Indica claramente se você deve **Repor Agora**, **Aguardar** ou se o estoque está **Saudável**.
        - **Sugestão de Envio:** Quantidade exata a ser enviada para atingir a cobertura desejada.
        - **Filtros Inteligentes:** Busque por SKU ou filtre apenas os itens críticos.

        #### 🗓️ Planejador de Envios
        Ferramenta de simulação para o próximo lote:
        - **Calculadora de Reposição:** Ajuste a cobertura desejada e veja o impacto imediato nas quantidades.
        - **Lead Time:** Considere o tempo de transporte no cálculo da sugestão.
        - **Resumo do Envio:** Total de unidades e volume estimado para o próximo envio.

        #### 📥 Exportar
        Leve seus dados para onde precisar:
        - **Excel do Painel:** Gera uma planilha pronta com todas as ações recomendadas.
        - **Relatório Completo:** Exporta todos os dados processados, incluindo auditorias e histórico de movimentações.

        #### 📊 Movimentações
        Detalhes e auditoria das entradas e saídas do seu estoque:
        - **Divergências Identificadas:** Apresenta SKUs com diferenças significativas entre retiradas e vendas, indicando possíveis inconsistências.
        - **Histórico Completo:** Tabela detalhada de todas as movimentações, permitindo uma análise aprofundada de cada transação.
        - **Análise de Tendências:** Gráficos que mostram o volume de entradas e saídas ao longo do tempo.
        """)

    # 4. Explicação dos Cálculos e Métricas
    with st.expander("Explicação dos Cálculos e Métricas", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("#### 🔍 Auditoria")
            st.markdown("""
            Compara **Retiradas** com **Vendas**.
            - **Fórmula:** `Abs(Retiradas - Vendas)`
            - **Alerta:** Diferença > 5 un. E > 20% vol.
            """)
        with c2:
            st.markdown("#### ⚡ Velocidade")
            st.markdown("""
            Baseada no relatório de 30 dias.
            - **Fórmula:** `Venda Diária = Unidades / 30`
            - **Cobertura:** `Estoque / Venda Diária`
            """)
        with c3:
            st.markdown("#### 📦 Sugestão")
            st.markdown("""
            Quanto enviar para manter o estoque.
            - **Fórmula:** `(Venda Diária * Dias * (1 + % Cresc.)) - Estoque`
            """)

    # 4. Dicas e Boas Práticas
    with st.expander("Dicas e Boas Práticas", expanded=False):
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            st.markdown("""
            #### ⚙️ Parâmetros
            - **Calibração Sazonal:** Ajusta automaticamente os dias de cobertura baseado em datas comerciais (ex: Black Friday).
            - **Crescimento Esperado (%):** Define quanto você espera vender a mais. 
            - **Dias de Cobertura:** Quantos dias de estoque você deseja manter no Full.
            """)
        with col_s2:
            st.markdown("""
            #### 🎯 Oportunidades
            - **Mínimo de Unidades:** Filtra SKUs que vendem acima deste valor mas não estão no Full.
            - **Top N Resultados:** Limita a quantidade de sugestões exibidas.
            - **Excluir Canceladas:** Limpa os dados de vendas para considerar apenas pedidos reais.
            """)

# ============================================================
# Parsing: Relatório geral de estoque
# ============================================================

@st.cache_data(show_spinner=False)
def parse_stock_general(xlsx_bytes: bytes) -> Dict:
    xls = pd.ExcelFile(BytesIO(xlsx_bytes))
    raw = pd.read_excel(xls, sheet_name="Resumo", header=None)

    updated_at = ""
    for i in range(min(10, len(raw))):
        row_txt = " ".join([_clean_str(v) for v in raw.iloc[i].values if _clean_str(v)])
        if "Atualizado em" in row_txt:
            updated_at = row_txt
            break

    cap_small = ""
    cap_large = ""
    for i in range(len(raw)):
        row_txt = " ".join([_clean_str(v) for v in raw.iloc[i].values if _clean_str(v)])
        if "Capacidade utilizada" in row_txt:
            nums = re.findall(r"(\d+(?:[.,]\d+)?)\s*%", row_txt)
            if len(nums) >= 2:
                cap_small = nums[0]
                cap_large = nums[1]
            break

    # Localiza tabela resumo - busca a linha que contém "SKU" como header
    start_row = None
    for i in range(min(20, len(raw))):
        row = raw.iloc[i].astype(str)
        row_values = [str(v).lower().strip() for v in row.values if pd.notna(v) and str(v).strip()]
        if 'sku' in row_values:
            start_row = i
            break
    
    if start_row is None:
        start_row = _find_row_index(raw, "SKU")
    if start_row is None:
        start_row = _find_row_index(raw, "Produto")
    if start_row is None:
        start_row = 0

    headers = [raw.iloc[start_row, j] for j in range(raw.shape[1])]
    headers = _make_unique_columns(headers)

    df = raw.iloc[start_row + 1 :, :].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    
    # Remove colunas completamente vazias
    df = df.dropna(axis=1, how='all')
    
    # Remove linhas onde SKU é vazio ou contém headers duplicados
    if "SKU" in df.columns:
        df = df[df["SKU"].astype(str).str.strip() != ""].copy()
        df = df[~df["SKU"].astype(str).str.lower().isin(["sku", "código", "codigo"])].copy()

    return {
        "updated_at": updated_at,
        "cap_small": cap_small,
        "cap_large": cap_large,
        "resumo_table": df,
    }


# ============================================================
# Parsing: Consolidado de movimentações
# ============================================================

@st.cache_data(show_spinner=False)
def parse_movements(xlsx_bytes: bytes, today: date) -> pd.DataFrame:
    xls = pd.ExcelFile(BytesIO(xlsx_bytes))
    raw = pd.read_excel(xls, sheet_name=0)
    
    # Se o DataFrame estiver vazio, retorna
    if raw.empty:
        return raw
        
    # Tenta encontrar a linha de cabeçalho (onde está o SKU ou Código)
    start_row = 0
    for i in range(min(10, len(raw))):
        row_values = [str(v).lower() for v in raw.iloc[i].values]
        if any(p in row_values for p in ["sku", "código", "codigo"]):
            start_row = i
            break
            
    # Se encontrou uma linha de cabeçalho que não é a primeira
    if start_row > 0 or any(p in [str(c).lower() for c in raw.columns] for p in ["sku", "código", "codigo"]):
        # Se o cabeçalho já estava nas colunas do pandas
        if any(p in [str(c).lower() for c in raw.columns] for p in ["sku", "código", "codigo"]):
            df = raw
        else:
            headers = [str(raw.iloc[start_row, j]) for j in range(raw.shape[1])]
            df = raw.iloc[start_row + 1 :, :].copy()
            df.columns = headers
    else:
        df = raw
        
    return df.reset_index(drop=True)


def summarize_movements(mov: pd.DataFrame) -> Dict:
    if mov is None or mov.empty:
        return {"start": 0, "end": 0, "retiradas_abs": 0, "vendidas_abs": 0}
    
    col_retiradas = _detect_column(mov.columns, ["retiradas", "retirada"])
    col_vendidas = _detect_column(mov.columns, ["vendidas", "vendida", "vendas"])
    
    retiradas = 0
    vendidas = 0
    
    if col_retiradas:
        retiradas = abs(mov[col_retiradas].apply(_to_number).fillna(0).sum())
    if col_vendidas:
        vendidas = abs(mov[col_vendidas].apply(_to_number).fillna(0).sum())
    
    return {
        "start": 0,
        "end": 0,
        "retiradas_abs": int(retiradas),
        "vendidas_abs": int(vendidas),
        "period_start": None,
        "period_end": None,
    }


def reconcile_audit(mov: pd.DataFrame, resumo: pd.DataFrame = None) -> pd.DataFrame:
    """Cria DataFrame de auditoria comparando movimentacoes com estoque."""
    if mov is None or mov.empty:
        return pd.DataFrame()
    
    audit = mov.copy()
    
    # Detecta coluna SKU de forma flexível
    col_sku = _detect_column(audit.columns, ["sku", "código", "codigo"])
    if col_sku:
        audit["SKU"] = audit[col_sku].astype(str).apply(_clean_str)
    else:
        # Se não encontrar SKU, tenta usar a primeira coluna como fallback ou levanta erro amigável
        if not audit.empty:
            audit["SKU"] = audit.iloc[:, 0].astype(str).apply(_clean_str)
        else:
            audit["SKU"] = ""
    
    # Detecta colunas de retiradas e vendidas
    col_retiradas = _detect_column(mov.columns, ["retiradas", "retirada"])
    col_vendidas = _detect_column(mov.columns, ["vendidas", "vendida", "vendas"])
    
    if col_retiradas:
        audit["_retiradas"] = audit[col_retiradas].apply(_to_number).fillna(0).abs()
    else:
        audit["_retiradas"] = 0
    
    if col_vendidas:
        audit["_vendidas"] = audit[col_vendidas].apply(_to_number).fillna(0).abs()
    else:
        audit["_vendidas"] = 0
    
    # Divergencia: retiradas muito diferentes de vendidas (mais de 20% de diferenca)
    audit["_diff"] = abs(audit["_retiradas"] - audit["_vendidas"])
    audit["_max_val"] = audit[["_retiradas", "_vendidas"]].max(axis=1)
    audit["_pct_diff"] = np.where(
        audit["_max_val"] > 0,
        audit["_diff"] / audit["_max_val"] * 100,
        0
    )
    
    # Considera divergencia se:
    # 1. Diferenca absoluta > 5 unidades E diferenca percentual > 20%
    # 2. Ou se tem retiradas mas zero vendas (ou vice-versa) com valores significativos
    audit["_ok"] = ~(
        ((audit["_diff"] > 5) & (audit["_pct_diff"] > 20)) |
        ((audit["_retiradas"] > 10) & (audit["_vendidas"] == 0)) |
        ((audit["_vendidas"] > 10) & (audit["_retiradas"] == 0))
    )
    
    return audit


# ============================================================
# Parsing: Relatório de vendas
# ============================================================

@st.cache_data(show_spinner=False)
def parse_sales_report(xlsx_bytes: bytes) -> Tuple[pd.DataFrame, Dict]:
    xls = pd.ExcelFile(BytesIO(xlsx_bytes))
    sh = xls.sheet_names[0]
    raw = pd.read_excel(xls, sheet_name=sh, header=None)

    header_row = None
    for i in range(min(60, len(raw))):
        row = raw.iloc[i].astype(str)
        if row.str.contains("SKU", case=False, na=False).any():
            header_row = i
            break

    if header_row is None:
        for i in range(min(60, len(raw))):
            row = raw.iloc[i].astype(str)
            if row.str.contains("Produto", case=False, na=False).any():
                header_row = i
                break

    if header_row is None:
        header_row = 0

    headers = [raw.iloc[header_row, j] for j in range(raw.shape[1])]
    headers = _make_unique_columns(headers)

    df = raw.iloc[header_row + 1 :, :].copy()
    df.columns = headers
    df = df.reset_index(drop=True)

    # Detecta colunas - atualizado para nomes exatos do relatório ML
    col_map = {
        "sku": _detect_column(df.columns, ["sku", "código", "codigo"]),
        "ad_id": _detect_column(df.columns, [
            "# de anúncio", "id do anúncio", "id anuncio", "anúncio", "anuncio"
        ]),
        "qtd": _detect_column(df.columns, ["unidades", "quantidade", "qtd"]),
        "valor": _detect_column(df.columns, [
            "receita por produtos (brl)", "receita por produtos", 
            "receita", "valor", "total (brl)", "total", "faturamento"
        ]),
        "produto": _detect_column(df.columns, [
            "título do anúncio", "titulo do anuncio", 
            "produto", "título", "titulo", "descrição"
        ]),
        "data_venda": _detect_column(df.columns, [
            "data da venda", "data venda", "data"
        ]),
        "status": _detect_column(df.columns, [
            "estado", "status", "situação", "situacao"
        ]),
        "variacao": _detect_column(df.columns, ["variação", "variacao"]),
        "sheet": sh,
        "header_auto": True,
    }

    return df, col_map


def compute_last_sale_by_sku(sales_raw: pd.DataFrame, sales_map: Dict, today: date) -> pd.DataFrame:
    """Calcula última venda por SKU."""
    if sales_raw is None or sales_raw.empty:
        return pd.DataFrame()
    
    sku_col = sales_map.get("sku")
    date_col = sales_map.get("data_venda")
    
    if not sku_col or sku_col not in sales_raw.columns:
        return pd.DataFrame()
    
    df = sales_raw.copy()
    df["SKU"] = df[sku_col].astype(str).apply(_clean_str)
    
    if date_col and date_col in df.columns:
        df["_date"] = pd.to_datetime(df[date_col], errors="coerce")
        grouped = df.groupby("SKU")["_date"].max().reset_index()
        grouped.columns = ["SKU", "last_sale_date"]
        grouped["days_since_last_sale"] = (pd.Timestamp(today) - grouped["last_sale_date"]).dt.days
        return grouped
    
    return pd.DataFrame(columns=["SKU", "last_sale_date", "days_since_last_sale"])


def build_sales_opportunities(
    sales_raw: pd.DataFrame,
    sales_map: Dict,
    stock: pd.DataFrame,
    min_sales_qty: float,
    top_n: int,
    exclude_canceled: bool,
) -> pd.DataFrame:
    """Identifica oportunidades de produtos fora do Full."""
    if sales_raw is None or sales_raw.empty:
        return pd.DataFrame()
    
    sku_col = sales_map.get("sku")
    qtd_col = sales_map.get("qtd")
    valor_col = sales_map.get("valor")
    prod_col = sales_map.get("produto")
    status_col = sales_map.get("status")
    ad_col = sales_map.get("ad_id")
    
    df = sales_raw.copy()
    
    # Filtra canceladas se necessário
    if exclude_canceled and status_col and status_col in df.columns:
        df = df[~df[status_col].astype(str).str.lower().str.contains("cancel|devolv", na=False)]
    
    # Prepara colunas
    if sku_col and sku_col in df.columns:
        df["SKU"] = df[sku_col].astype(str).apply(_clean_str)
    else:
        df["SKU"] = ""
    
    if ad_col and ad_col in df.columns:
        df["ID_Anuncio"] = df[ad_col].astype(str).apply(_clean_str)
    else:
        df["ID_Anuncio"] = ""
    
    if prod_col and prod_col in df.columns:
        df["Produto"] = df[prod_col].astype(str).apply(_clean_str)
    else:
        df["Produto"] = ""
    
    if qtd_col and qtd_col in df.columns:
        df["vendas_periodo"] = df[qtd_col].apply(_to_number).fillna(0)
    else:
        df["vendas_periodo"] = 0
    
    if valor_col and valor_col in df.columns:
        df["receita_produtos"] = df[valor_col].apply(_to_number).fillna(0)
        df["total_brl"] = df["receita_produtos"]
    else:
        df["receita_produtos"] = 0
        df["total_brl"] = 0
    
    # Agrupa por SKU
    key_col = "SKU" if df["SKU"].str.strip().ne("").any() else "ID_Anuncio"
    
    agg = df.groupby(key_col).agg({
        "Produto": "first",
        "vendas_periodo": "sum",
        "total_brl": "sum"
    }).reset_index()
    
    # Filtra SKUs que já estão no Full
    full_skus = set()
    if stock is not None and not stock.empty:
        if "SKU" in stock.columns:
            full_skus = set(stock["SKU"].astype(str).apply(_clean_str))
    
    agg = agg[~agg[key_col].isin(full_skus)]
    
    # Filtra por mínimo de vendas
    agg = agg[agg["vendas_periodo"] >= min_sales_qty]
    
    # Ordena e limita
    agg = agg.sort_values("vendas_periodo", ascending=False).head(top_n)
    
    return agg


def add_suggestion_to_sales_opps(opps: pd.DataFrame, cover_days: int, growth_pct: float) -> pd.DataFrame:
    if opps is None or opps.empty:
        return pd.DataFrame()
    
    df = opps.copy()
    # Assume que o período de vendas é 30 dias
    df["Velocidade Média Dia"] = df["vendas_periodo"] / 30
    df["Sugestão Envio (un.)"] = np.ceil(df["Velocidade Média Dia"] * cover_days * (1 + growth_pct))
    
    return df


# ============================================================
# Painel de Ação
# ============================================================

def build_action_panel(
    resumo: pd.DataFrame,
    mov: pd.DataFrame,
    last_sale: pd.DataFrame,
    cover_days: int,
    growth_pct: float
) -> pd.DataFrame:
    if resumo is None or resumo.empty:
        return pd.DataFrame()
    
    df = resumo.copy()
    df["SKU"] = df["SKU"].astype(str).apply(_clean_str)
    
    # Detecta colunas essenciais - atualizado para nomes exatos do relatório ML
    col_estoque = _detect_column(df.columns, [
        "unidades no full", "aptas para venda", "full - aptas para venda", 
        "estoque", "disponível", "unidades que ocupan espacio"
    ])
    col_vendas = _detect_column(df.columns, [
        "vendas últimos 30 dias (un.)", "vendas últimos 30 dias", 
        "vendas 30d", "vendas", "velocidade de vendas"
    ])
    col_a_caminho = _detect_column(df.columns, [
        "unidades a caminho do full", "a caminho", "em trânsito", 
        "transito", "unidades a caminho"
    ])
    col_tempo = _detect_column(df.columns, [
        "tempo até esgotar estoque", "tempo até esgotar", "dias de estoque"
    ])
    col_produto = _detect_column(df.columns, ["produto", "título", "titulo", "descrição"])
    
    # Prepara dados numéricos
    df["Estoque Atual"] = df[col_estoque].apply(_to_number).fillna(0) if col_estoque else 0
    df["Vendas Período"] = df[col_vendas].apply(_to_number).fillna(0) if col_vendas else 0
    df["A caminho"] = df[col_a_caminho].apply(_to_number).fillna(0) if col_a_caminho else 0
    
    # Garante que a coluna Produto existe e está preenchida
    if col_produto and col_produto in df.columns:
        df["Produto"] = df[col_produto].astype(str).apply(_clean_str)
    elif "Produto" not in df.columns:
        df["Produto"] = ""
    
    # Velocidade e Sugestão
    df["Velocidade Média Dia"] = df["Vendas Período"] / 30
    df["Sugestão Envio (un.)"] = np.ceil(
        (df["Velocidade Média Dia"] * cover_days * (1 + growth_pct)) - df["Estoque Atual"] - df["A caminho"]
    )
    df["Sugestão Envio (un.)"] = np.maximum(0, df["Sugestão Envio (un.)"])
    
    # Ação Recomendada
    def get_action(row):
        sug = row["Sugestão Envio (un.)"]
        est = row["Estoque Atual"]
        vendas = row["Vendas Período"]
        
        if est <= 0 and vendas > 0:
            return "🚨 REPOR AGORA"
        if sug > 0:
            return "🛑 REPOR EM BREVE"
        if est > 0 and vendas > 0:
            return "📢 MANTER"
        return "⚪ SEM GIRO"

    df["Ação Recomendada"] = df.apply(get_action, axis=1)
    
    # Prioridade para ordenação
    prio_map = {"🚨 REPOR AGORA": 0, "🛑 REPOR EM BREVE": 1, "📢 MANTER": 2, "⚪ SEM GIRO": 3}
    df["prioridade"] = df["Ação Recomendada"].map(prio_map)
    
    # Merge com última venda
    if last_sale is not None and not last_sale.empty:
        df = df.merge(last_sale[["SKU", "days_since_last_sale"]], on="SKU", how="left")
    
    return df


# ============================================================
# UI - Premium Design
# ============================================================

today_sp = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
season = seasonal_calibration(today_sp)

# ===== SIDEBAR PREMIUM =====
with st.sidebar:
    # Logo Area
    st.markdown("""
    <div class="logo-area">
        <div class="logo-text">Full Dashboard</div>
        <div class="logo-subtitle">Gestão Inteligente de Estoque</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Arquivos Section
    st.markdown('<div class="sidebar-section-title">Arquivos</div>', unsafe_allow_html=True)
    stock_file = st.file_uploader("Relatório geral de estoque", type=["xlsx"], key="stock", help="Arquivo Excel com o relatório geral de estoque do Mercado Livre Full")

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Movimentações Section
    st.markdown('<div class="sidebar-section-title">Movimentações</div>', unsafe_allow_html=True)
    use_mov = st.checkbox("Ativar modo completo", value=False, help="Habilita auditoria e leitura avançada de movimentações")
    mov_file = st.file_uploader("Consolidado de movimentações", type=["xlsx"], key="mov", help="Arquivo opcional com consolidado de movimentações") if use_mov else None

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Vendas Section
    st.markdown('<div class="sidebar-section-title">Vendas</div>', unsafe_allow_html=True)
    sales_file = st.file_uploader("Relatório de vendas 30d", type=["xlsx"], key="sales", help="Arquivo opcional com relatório de vendas dos últimos 30 dias")

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Parâmetros Section
    st.markdown('<div class="sidebar-section-title">Parâmetros</div>', unsafe_allow_html=True)
    use_seasonal = st.checkbox("Calibração sazonal automática", value=True, help="Ajusta automaticamente a cobertura baseado em datas comerciais")
    growth_pct = st.number_input("Crescimento esperado (%)", 0.0, 300.0, 20.0, 5.0, help="Percentual de crescimento esperado para o período") / 100.0
    target_cover_days_manual = st.number_input("Dias de cobertura", 1, 120, int(season["cover_days"]), 5, help="Número de dias de cobertura de estoque desejado")
    cover_days_effective = int(season["cover_days"]) if use_seasonal else int(target_cover_days_manual)

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Oportunidades Section
    st.markdown('<div class="sidebar-section-title">Oportunidades</div>', unsafe_allow_html=True)
    min_sales_qty = st.number_input("Mínimo de unidades", 1.0, 10000.0, 5.0, 1.0, help="Quantidade mínima de vendas para considerar como oportunidade")
    top_n = st.number_input("Top N resultados", 5, 500, 50, 5, help="Número máximo de oportunidades a exibir")
    exclude_canceled_and_returns = st.checkbox("Excluir canceladas", value=True, help="Remove pedidos cancelados e devoluções da análise")

    params = {"growth_pct": float(growth_pct)}

    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Simulação Section
    st.markdown('<div class="sidebar-section-title">Simulação A/B</div>', unsafe_allow_html=True)
    enable_ab = st.checkbox("Habilitar simulação", value=True, help="Mostra comparativo entre cenários com e sem sazonalidade")
    next_ship_date = st.date_input("Data do envio", value=today_sp + timedelta(days=7), disabled=not enable_ab, help="Data prevista para o próximo envio")
    lead_time_days = st.number_input("Lead time (dias)", 0, 120, 20, 1, disabled=not enable_ab, help="Tempo mínimo necessário antes de um evento")
    cover_days_no_season = st.number_input("Cobertura sem sazonalidade", 1, 120, int(target_cover_days_manual), 5, disabled=not enable_ab, help="Dias de cobertura para cenário sem sazonalidade")

# ===== MAIN CONTENT =====
if not stock_file:
    # Header
    st.markdown("""
    <div class="premium-header">
        <h1>Dashboard Fulfillment Estratégico</h1>
        <p class="premium-subtitle">Gestão Inteligente de Estoque</p>
    </div>
    """, unsafe_allow_html=True)

    # Tabs for Welcome Screen
    tw1, tw2 = st.tabs(["Início", "Guia de Uso"])

    with tw1:
        st.markdown("""
        <div class="info-box" style="text-align: center; max-width: 600px; margin: 2rem auto;">
            <p style="font-size: 1.1rem; margin-bottom: 0.5rem;">👋 <strong>Bem-vindo ao Full Dashboard Pro</strong></p>
            <p style="margin: 0;">Faça upload do <strong>Relatório geral de estoque</strong> na barra lateral para começar a análise.</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Features Grid
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("""
            <div class="glass-card" style="text-align: center;">
                <div class="icon-box" style="margin: 0 auto 1rem auto; width: 48px; height: 48px; border-radius: 12px; background: rgba(46, 125, 50, 0.2); border: 1px solid var(--primary);">
                    <img src="https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/trending-up.svg" style="width: 24px; height: 24px; filter: brightness(0) invert(1);">
                </div>
                <h3 style="margin-bottom: 0.5rem;">Análise Inteligente</h3>
                <p style="font-size: 0.875rem;">Sugestões automáticas de reposição baseadas em velocidade de vendas e sazonalidade.</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("""
            <div class="glass-card" style="text-align: center;">
                <div class="icon-box" style="margin: 0 auto 1rem auto; width: 48px; height: 48px; border-radius: 12px; background: rgba(46, 125, 50, 0.2); border: 1px solid var(--primary);">
                    <img src="https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/target.svg" style="width: 24px; height: 24px; filter: brightness(0) invert(1);">
                </div>
                <h3 style="margin-bottom: 0.5rem;">Oportunidades</h3>
                <p style="font-size: 0.875rem;">Identifique produtos com alto potencial que ainda não estão no Full.</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown("""
            <div class="glass-card" style="text-align: center;">
                <div class="icon-box" style="margin: 0 auto 1rem auto; width: 48px; height: 48px; border-radius: 12px; background: rgba(46, 125, 50, 0.2); border: 1px solid var(--primary);">
                    <img src="https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/calendar.svg" style="width: 24px; height: 24px; filter: brightness(0) invert(1);">
                </div>
                <h3 style="margin-bottom: 0.5rem;">Calendário Sazonal</h3>
                <p style="font-size: 0.875rem;">Ajuste automático para datas comerciais importantes do varejo brasileiro.</p>
            </div>
            """, unsafe_allow_html=True)
    
    with tw2:
        show_guia_uso()
    
    st.stop()

# Parse estoque
try:
    stock = parse_stock_general(stock_file.getvalue())
except Exception as e:
    st.error(f"Não consegui ler o relatório geral de estoque. Erro: {e}")
    st.stop()

# Parse movimentações (se houver)
mov: Optional[pd.DataFrame] = None
mov_summary: Optional[Dict] = None
audit = pd.DataFrame()

if mov_file is not None:
    try:
        mov = parse_movements(mov_file.getvalue(), today_sp)
        mov_summary = summarize_movements(mov)
        audit = reconcile_audit(mov)
    except Exception as e:
        st.warning(f"Não consegui ler o consolidado de movimentações. Vou seguir em modo leve. Erro: {e}")
        mov = None
        mov_summary = None
        audit = pd.DataFrame()

resumo_tbl = stock["resumo_table"]

# ===== HEADER =====
st.markdown("""
<div class="premium-header">
    <h1>Dashboard Fulfillment Estratégico</h1>
    <p class="premium-subtitle">Gestão Inteligente de Estoque</p>
</div>
""", unsafe_allow_html=True)

# Context Bar
mode_txt = "Modo Completo" if mov is not None else "Modo Leve"
mode_icon_url = "https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/zap.svg"
season_icon_url = "https://raw.githubusercontent.com/lucide-icons/lucide/main/icons/calendar.svg"

st.markdown(f"""
<div class="context-bar">
    <div class="context-item">
        <span class="icon-box"><img src="{mode_icon_url}" style="filter: brightness(0) invert(1); width: 16px; height: 16px;"></span>
        <span class="value">{mode_txt}</span>
    </div>
    <div class="context-item">
        <span class="label">Cobertura:</span>
        <span class="value">{int(cover_days_effective)} dias</span>
    </div>
    <div class="context-item">
        <span class="label">Data:</span>
        <span class="value">{today_sp.strftime('%d/%m/%Y')}</span>
    </div>
    <div class="context-item">
        <span class="icon-box"><img src="{season_icon_url}" style="filter: brightness(0) invert(1); width: 16px; height: 16px;"></span>
        <span class="value">{season.get('mode', 'Padrão')}</span>
    </div>
</div>
""", unsafe_allow_html=True)

# Oportunidade escondida (estoque)
opps_full_report = pd.DataFrame()
if (
    not resumo_tbl.empty
    and "Vendas últimos 30 dias (un.)" in resumo_tbl.columns
    and "Full - Aptas para venda" in resumo_tbl.columns
):
    tmp = resumo_tbl.copy()
    tmp["Vendas últimos 30 dias (un.)"] = tmp["Vendas últimos 30 dias (un.)"].map(_to_number).fillna(0.0)
    tmp["Full - Aptas para venda"] = tmp["Full - Aptas para venda"].map(_to_number).fillna(0.0)
    opps_full_report = tmp[(tmp["Vendas últimos 30 dias (un.)"] > 0) & (tmp["Full - Aptas para venda"] == 0)].copy().reset_index(drop=True)

# Vendas 30d opcional
sales_opps = pd.DataFrame()
last_sale_df = pd.DataFrame()
sales_map = None
sales_raw = None

if sales_file:
    try:
        sales_raw, sales_map = parse_sales_report(sales_file.getvalue())
    except Exception as e:
        st.warning(f"Não consegui ler o relatório de vendas opcional. Erro: {e}")
        sales_raw, sales_map = None, None

# Ajuste opcional de mapeamento
if sales_raw is not None and sales_map is not None:
    with st.expander("🔧 Ajustar mapeamento de colunas", expanded=False):
        cols_list = list(sales_raw.columns)

        c1, c2, c3 = st.columns(3)
        sku_col = c1.selectbox("Coluna SKU", options=cols_list, index=cols_list.index(sales_map["sku"]) if sales_map.get("sku") in cols_list else 0)
        ad_col = c2.selectbox("Coluna ID do anúncio", options=cols_list, index=cols_list.index(sales_map["ad_id"]) if sales_map.get("ad_id") in cols_list else 0)
        qtd_col = c3.selectbox("Coluna Quantidade", options=cols_list, index=cols_list.index(sales_map["qtd"]) if sales_map.get("qtd") in cols_list else 0)

        c4, c5, c6 = st.columns(3)
        val_col = c4.selectbox("Coluna Receita", options=cols_list, index=cols_list.index(sales_map["valor"]) if sales_map.get("valor") in cols_list else 0)
        prod_col = c5.selectbox("Coluna Produto", options=cols_list, index=cols_list.index(sales_map["produto"]) if sales_map.get("produto") in cols_list else 0)
        date_col = c6.selectbox("Coluna Data", options=cols_list, index=cols_list.index(sales_map["data_venda"]) if sales_map.get("data_venda") in cols_list else 0)

        sales_map["sku"] = sku_col
        sales_map["ad_id"] = ad_col
        sales_map["qtd"] = qtd_col
        sales_map["valor"] = val_col
        sales_map["produto"] = prod_col
        sales_map["data_venda"] = date_col

if sales_raw is not None and sales_map is not None:
    last_sale_df = compute_last_sale_by_sku(sales_raw, sales_map, today_sp)
    sales_opps = build_sales_opportunities(
        sales_raw, sales_map, resumo_tbl, float(min_sales_qty), int(top_n), bool(exclude_canceled_and_returns)
    )
    sales_opps = add_suggestion_to_sales_opps(sales_opps, cover_days_effective, float(growth_pct))

# Painel de ação
action_panel = build_action_panel(resumo_tbl, mov, last_sale_df, cover_days_effective, float(growth_pct))

# ===== TABS =====
if mov is not None:
    t1, t6, t2, t3, t4, t7, t5, t8 = st.tabs([
        "Resumo",
        "Análise Estratégica",
        "Oportunidades",
        "Movimentações",
        "Painel de Ação",
        "Planejador de Envios",
        "Exportar",
        "Guia de Uso"
    ])
else:
    t1, t6, t2, t4, t7, t5, t8 = st.tabs([
        "Resumo",
        "Análise Estratégica",
        "Oportunidades",
        "Painel de Ação",
        "Planejador de Envios",
        "Exportar",
        "Guia de Uso"
    ])
    t3 = None

# ============================================================
# TAB 1: Resumo
# ============================================================

with t1:
    # KPIs principais
    st.markdown("### Visão Geral")
    
    ap = action_panel.copy()
    ap["__estoque"] = pd.to_numeric(ap.get("Estoque Atual", 0), errors="coerce").fillna(0.0)
    ap["__vendas"] = pd.to_numeric(ap.get("Vendas Período", 0), errors="coerce").fillna(0.0)
    ap["__sug"] = pd.to_numeric(ap.get("Sugestão Envio (un.)", 0), errors="coerce").fillna(0.0)
    
    total_skus = len(ap)
    total_estoque = int(ap["__estoque"].sum())
    total_vendas = int(ap["__vendas"].sum())
    total_sugestao = int(ap["__sug"].sum())
    repor_count = int((ap["Ação Recomendada"] == "🚨 REPOR AGORA").sum())
    
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Total SKUs", _format_int(total_skus))
    k2.metric("Estoque Total", _format_int(total_estoque))
    k3.metric("Vendas Período", _format_int(total_vendas))
    k4.metric("Sugestão Total", _format_int(total_sugestao))
    k5.metric("Repor Agora", str(repor_count), delta=None if repor_count == 0 else f"{repor_count} urgentes", delta_color="inverse")
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Capacidade
    if stock.get("cap_small") or stock.get("cap_large"):
        st.markdown("### Capacidade do Armazém")
        cap1, cap2 = st.columns(2)
        if stock.get("cap_small"):
            cap1.metric("Pequeno/Médio", f"{stock['cap_small']}%")
        if stock.get("cap_large"):
            cap2.metric("Grande", f"{stock['cap_large']}%")
        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Contexto sazonal e Simulação A/B
    st.markdown("### Contexto Sazonal e Simulação A/B")
    
    col_season1, col_season2 = st.columns([2, 1])
    
    with col_season1:
        st.markdown(f"""
        <div class="glass-card">
            <p style="color: var(--text-muted); font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.5rem;">Modo Atual</p>
            <p style="color: var(--text-primary); font-size: 1.25rem; font-weight: 600; margin-bottom: 1rem;">{season.get('mode', 'Padrão')}</p>
            <p style="color: var(--text-secondary); font-size: 0.9rem; margin: 0;">{season.get('context', '')}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if enable_ab:
            st.markdown("#### 🔬 Simulação de Impacto A/B")
            # Cálculo rápido de impacto
            sug_com = total_sugestao
            
            # Garantir que __vel existe para a simulação
            if "__vel" not in ap.columns:
                ap["__vel"] = pd.to_numeric(ap.get("Velocidade Média Dia", 0), errors="coerce").fillna(0.0)
            
            # Simulação sem sazonalidade
            vel_media = ap["__vel"].mean() if not ap.empty else 0
            sug_sem = int(max(0, (vel_media * cover_days_no_season * total_skus) - total_estoque))
            diff_sug = sug_com - sug_sem
            
            c_ab1, c_ab2 = st.columns(2)
            with c_ab1:
                st.markdown(f"""
                <div class="health-card info">
                    <div class="health-label">Cenário Sazonal</div>
                    <div class="health-value">{sug_com:,}</div>
                    <div class="health-subtitle">unidades sugeridas</div>
                </div>
                """, unsafe_allow_html=True)
            with c_ab2:
                st.markdown(f"""
                <div class="health-card warning">
                    <div class="health-label">Cenário Padrão</div>
                    <div class="health-value">{sug_sem:,}</div>
                    <div class="health-subtitle">unidades sugeridas</div>
                </div>
                """, unsafe_allow_html=True)
            
            if diff_sug > 0:
                st.info(f"💡 A calibração sazonal sugere o envio de **{diff_sug:,} unidades extras** para garantir a cobertura no período de alta.")
            elif diff_sug < 0:
                st.success(f"✅ A calibração sazonal permite reduzir o envio em **{abs(diff_sug):,} unidades** sem comprometer a cobertura.")
            
            # Tabela Comparativa Detalhada
            st.markdown("#### 📊 Comparativo Detalhado por SKU")
            comp_df = ap.copy()
            comp_df["Sugestão (Sazonal)"] = comp_df["__sug"]
            comp_df["Sugestão (Padrão)"] = np.ceil((comp_df["__vel"] * cover_days_no_season) - comp_df["__estoque"] - comp_df.get("__a_caminho", 0))
            comp_df["Sugestão (Padrão)"] = comp_df["Sugestão (Padrão)"].clip(lower=0)
            comp_df["Diferença"] = comp_df["Sugestão (Sazonal)"] - comp_df["Sugestão (Padrão)"]
            
            show_cols = ["SKU", "Produto", "Sugestão (Sazonal)", "Sugestão (Padrão)", "Diferença"]
            st.dataframe(comp_df[comp_df["Diferença"] != 0][show_cols].sort_values("Diferença", ascending=False), use_container_width=True)
    
    with col_season2:
        # Sempre mostrar o próximo evento sazonal
        next_event_name = season.get('next_event', 'Sem eventos próximos')
        days_to_event = season.get('days_to_event')
        priority = season.get('priority', 'NORMAL')
        
        # Cor baseada na prioridade do evento
        priority_colors = {
            "CRITICO": "var(--danger)",
            "ALTO": "var(--warning)",
            "MEDIO": "var(--accent)",
            "NORMAL": "var(--primary-light)",
            "BAIXO": "var(--text-secondary)",
        }
        event_color = priority_colors.get(priority, "var(--primary-light)")
        
        st.markdown(f"""
        <div class="glass-card" style="text-align: center;">
            <p style="color: var(--text-muted); font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.5rem;">Próximo Evento</p>
            <p style="color: {event_color}; font-size: 1.25rem; font-weight: 700; margin-bottom: 0.25rem;">{next_event_name}</p>
            <p style="color: var(--text-secondary); font-size: 0.9rem; margin: 0;">em {days_to_event if days_to_event else '---'} dias</p>
            <p style="color: {event_color}; font-size: 0.75rem; font-weight: 600; margin-top: 0.5rem; text-transform: uppercase;">Prioridade: {priority}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if enable_ab:
            # Status do Time de Envio (Lead Time)
            days_to_compare = int(days_to_event) if days_to_event is not None else 999
            in_time = days_to_compare >= lead_time_days
            
            if in_time:
                status_color = "var(--primary-light)"
                status_text = "DENTRO DO PRAZO"
                status_icon = "✅"
            else:
                status_color = "var(--danger)"
                status_text = "PRAZO CRÍTICO"
                status_icon = "🚨"
            
            # Calcular data limite para envio
            if days_to_event:
                from datetime import timedelta as td
                data_limite = today_sp + td(days=days_to_event - lead_time_days)
                data_limite_str = data_limite.strftime('%d/%m') if days_to_event > lead_time_days else "URGENTE"
            else:
                data_limite_str = "---"
            
            st.markdown(f"""
            <div class="glass-card" style="text-align: center; margin-top: 1rem;">
                <p style="color: var(--text-muted); font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.5rem;">Status de Envio</p>
                <p style="color: {status_color}; font-size: 1.25rem; font-weight: 800; margin-bottom: 0.5rem;">{status_icon} {status_text}</p>
                <p style="color: var(--text-secondary); font-size: 0.85rem; margin: 0;">Lead Time: {lead_time_days} dias</p>
                <p style="color: var(--text-secondary); font-size: 0.85rem; margin: 0;">Data Limite Envio: {data_limite_str}</p>
                <p style="color: var(--text-secondary); font-size: 0.85rem; margin: 0;">Data Prevista: {next_ship_date.strftime('%d/%m')}</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Alerta se estiver fora do prazo
            if not in_time:
                st.error(f"🚨 **ATENÇÃO:** Você está fora do prazo para o evento {next_event_name}! O lead time de {lead_time_days} dias é maior que os {days_to_event} dias restantes.")
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # ===== PAINEL DE SAZONALIDADE DE MODA =====
    # Detectar se o relatório contém produtos de moda/vestuário
    fashion_detection = detect_fashion_category(resumo_tbl)
    
    if fashion_detection["is_fashion"]:
        st.markdown("### 👗 Sazonalidade de Vestuário por Estação")
        
        # Obter estação atual
        current_season = get_current_fashion_season(today_sp)
        
        # Layout principal: Estação atual + Cards das estações lado a lado
        main_col1, main_col2 = st.columns([1, 2])
        
        with main_col1:
            # Card da estação atual em destaque
            st.markdown(f"""
            <div class="glass-card" style="text-align: center; border: 3px solid {current_season['cor_tema']}; padding: 1.5rem;">
                <p style="font-size: 3rem; margin: 0;">{current_season['emoji']}</p>
                <p style="color: {current_season['cor_tema']}; font-size: 1.5rem; font-weight: 800; margin: 0.5rem 0;">{current_season['nome']}</p>
                <p style="color: var(--accent); font-size: 0.85rem; font-weight: 600; margin: 0.5rem 0;">ESTAÇÃO ATUAL</p>
                <div style="margin-top: 1rem; padding-top: 1rem; border-top: 1px solid rgba(255,255,255,0.1);">
                    <p style="color: {current_season['cor_tema']}; font-size: 2.5rem; font-weight: 800; margin: 0;">{current_season['dias_restantes']}</p>
                    <p style="color: var(--text-muted); font-size: 0.8rem; margin: 0;">dias restantes</p>
                </div>
                <div style="margin-top: 1rem; padding-top: 1rem; border-top: 1px solid rgba(255,255,255,0.1);">
                    <p style="color: var(--text-muted); font-size: 0.7rem; text-transform: uppercase; margin: 0;">Próxima</p>
                    <p style="color: var(--text-primary); font-size: 1rem; font-weight: 600; margin: 0.25rem 0;">{current_season['proxima_estacao']}</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with main_col2:
            # Mini cards das 4 estações em linha
            st.markdown("""
            <div style="display: flex; gap: 0.5rem; margin-bottom: 1rem;">
            """, unsafe_allow_html=True)
            
            for key, season_data in FASHION_SEASONS.items():
                is_current = key == current_season["key"]
                border = f"2px solid {season_data['cor_tema']}" if is_current else "1px solid rgba(255,255,255,0.1)"
                opacity = "1" if is_current else "0.5"
                bg = f"rgba(255,255,255,0.05)" if is_current else "transparent"
                
                st.markdown(f"""
                <div style="flex: 1; text-align: center; padding: 0.75rem; border-radius: 0.75rem; border: {border}; opacity: {opacity}; background: {bg};">
                    <p style="font-size: 1.25rem; margin: 0;">{season_data['emoji']}</p>
                    <p style="color: {season_data['cor_tema']}; font-size: 0.75rem; font-weight: 600; margin: 0.25rem 0;">{season_data['nome']}</p>
                    <p style="color: var(--text-muted); font-size: 0.6rem; margin: 0;">{season_data['inicio'][1]:02d}/{season_data['inicio'][0]:02d} - {season_data['fim'][1]:02d}/{season_data['fim'][0]:02d}</p>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Produtos em destaque - tags em linha horizontal
            produtos_tags = " ".join([f'<span style="background: rgba(255,255,255,0.1); padding: 0.35rem 0.75rem; border-radius: 1rem; font-size: 0.8rem; color: var(--text-primary); white-space: nowrap;">{p}</span>' for p in current_season['produtos_destaque']])
            
            st.markdown(f"""
            <div class="glass-card" style="padding: 1rem;">
                <p style="color: var(--text-muted); font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 0.75rem;">Produtos em Destaque para {current_season['nome']}</p>
                <div style="display: flex; flex-wrap: wrap; gap: 0.5rem;">
                    {produtos_tags}
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Recomendação de estoque compacta
            st.markdown(f"""
            <div style="background: rgba(76, 175, 80, 0.1); border-left: 3px solid var(--accent); padding: 0.75rem 1rem; border-radius: 0 0.5rem 0.5rem 0; margin-top: 0.75rem;">
                <p style="color: var(--text-primary); font-size: 0.85rem; margin: 0;"><strong>📦 Dica:</strong> {current_season['recomendacao_estoque']}</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Info de detecção compacta
        st.markdown(f"""
        <p style="color: var(--text-muted); font-size: 0.75rem; text-align: center; margin-top: 1rem;">
            👗 Categoria de moda detectada: {fashion_detection['fashion_products']} de {fashion_detection['total_products']} produtos ({fashion_detection['confidence']}%)
        </p>
        """, unsafe_allow_html=True)
        
        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # Tabela resumo
    st.markdown("### Dados do Estoque")
    show_df(resumo_tbl, height=500)

# ============================================================
# TAB 6: Análise Estratégica
# ============================================================

with t6:
    st.markdown("### Análise Estratégica do Estoque")
    
    # Preparar dados para análise
    ap = action_panel.copy()
    ap["__estoque"] = pd.to_numeric(ap.get("Estoque Atual", 0), errors="coerce").fillna(0.0)
    ap["__vendas"] = pd.to_numeric(ap.get("Vendas Período", ap.get("Vendas 30d", 0)), errors="coerce").fillna(0.0)
    ap["__vendas_30d"] = pd.to_numeric(ap.get("Vendas 30d", ap.get("Vendas Período", 0)), errors="coerce").fillna(0.0)
    ap["__sug"] = pd.to_numeric(ap.get("Sugestão Envio (un.)", 0), errors="coerce").fillna(0.0)
    ap["__vel"] = pd.to_numeric(ap.get("Velocidade Média Dia", 0), errors="coerce").fillna(0.0)
    ap["__a_caminho"] = pd.to_numeric(ap.get("A caminho", 0), errors="coerce").fillna(0.0)
    
    # Calcular dias de estoque
    ap["__dias_estoque"] = np.where(
        ap["__vel"] > 0,
        ap["__estoque"] / ap["__vel"],
        np.where(ap["__estoque"] > 0, 999, 0)
    )
    
    # Métricas de saúde
    total_skus = len(ap)
    skus_com_estoque = int((ap["__estoque"] > 0).sum())
    skus_sem_estoque = int((ap["__estoque"] <= 0).sum())
    skus_com_vendas = int((ap["__vendas_30d"] > 0).sum())
    
    # Rupturas (sem estoque mas com vendas)
    rupturas = int(((ap["__estoque"] <= 0) & (ap["__vendas_30d"] > 0)).sum())
    taxa_ruptura = (rupturas / skus_com_vendas * 100) if skus_com_vendas > 0 else 0
    
    # Peso morto (estoque sem vendas)
    peso_morto = int(((ap["__estoque"] > 0) & (ap["__vendas_30d"] <= 0)).sum())
    taxa_peso_morto = (peso_morto / skus_com_estoque * 100) if skus_com_estoque > 0 else 0
    
    # Cobertura média
    cobertura_media = ap.loc[ap["__vel"] > 0, "__dias_estoque"].mean() if (ap["__vel"] > 0).any() else 0
    
    # Estoque crítico (< 7 dias)
    estoque_critico = int(((ap["__dias_estoque"] > 0) & (ap["__dias_estoque"] < 7) & (ap["__vendas_30d"] > 0)).sum())
    
    # Saúde geral do estoque
    saude_score = 100 - (taxa_ruptura * 2) - (taxa_peso_morto * 0.5)
    saude_score = max(0, min(100, saude_score))
    
    if saude_score >= 80:
        saude_class = "good"
        saude_label = "Saudável"
    elif saude_score >= 60:
        saude_class = "warning"
        saude_label = "Atenção"
    else:
        saude_class = "critical"
        saude_label = "Crítico"
    
    # ===== CARDS DE SAÚDE =====
    st.markdown("#### Indicadores de Saúde")
    
    h1, h2, h3, h4 = st.columns(4)
    
    with h1:
        st.markdown(f"""
        <div class="health-card {saude_class}">
            <div class="health-label">Saúde do Estoque</div>
            <div class="health-value">{saude_score:.0f}%</div>
            <div class="health-subtitle">{saude_label}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with h2:
        ruptura_class = "critical" if taxa_ruptura > 15 else "warning" if taxa_ruptura > 5 else "good"
        st.markdown(f"""
        <div class="health-card {ruptura_class}">
            <div class="health-label">Taxa de Ruptura</div>
            <div class="health-value">{taxa_ruptura:.1f}%</div>
            <div class="health-subtitle">{rupturas} SKUs vendendo sem estoque</div>
        </div>
        """, unsafe_allow_html=True)
    
    with h3:
        peso_class = "critical" if taxa_peso_morto > 20 else "warning" if taxa_peso_morto > 10 else "good"
        st.markdown(f"""
        <div class="health-card {peso_class}">
            <div class="health-label">Peso Morto</div>
            <div class="health-value">{taxa_peso_morto:.1f}%</div>
            <div class="health-subtitle">{peso_morto} SKUs parados</div>
        </div>
        """, unsafe_allow_html=True)
    
    with h4:
        cob_class = "critical" if cobertura_media < 7 else "warning" if cobertura_media < 15 else "good"
        st.markdown(f"""
        <div class="health-card {cob_class}">
            <div class="health-label">Cobertura Média</div>
            <div class="health-value">{cobertura_media:.0f}</div>
            <div class="health-subtitle">dias de estoque</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # ===== ALERTAS IMPORTANTES =====
    st.markdown("#### Alertas Importantes")
    
    alertas_html = ""
    
    if rupturas > 0:
        alertas_html += f'<span class="alert-badge critical">🚨 {rupturas} SKUs em ruptura</span> '
    
    if estoque_critico > 0:
        alertas_html += f'<span class="alert-badge warning">⚠️ {estoque_critico} SKUs com estoque crítico (&lt;7 dias)</span> '
    
    if peso_morto > 0:
        alertas_html += f'<span class="alert-badge info">📦 {peso_morto} SKUs sem giro</span> '
    
    if not alertas_html:
        alertas_html = '<span class="alert-badge success">✅ Nenhum alerta crítico</span>'
    
    st.markdown(f'<div style="display: flex; flex-wrap: wrap; gap: 0.75rem; margin-bottom: 1rem;">{alertas_html}</div>', unsafe_allow_html=True)
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    st.markdown("#### Distribuição do Estoque")
    
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        # Gráfico de distribuição por ação recomendada
        st.markdown("##### Por Ação Recomendada")
        
        acao_counts = ap["Ação Recomendada"].value_counts()
        
        for acao, count in acao_counts.items():
            pct = count / total_skus * 100
            if "🚨" in str(acao):
                bar_class = "critical"
            elif "🛑" in str(acao):
                bar_class = "warning"
            elif "📢" in str(acao):
                bar_class = "good"
            else:
                bar_class = "good"
            
            st.markdown(f"""
            <div style="margin-bottom: 0.75rem;">
                <div style="display: flex; justify-content: space-between; margin-bottom: 0.25rem;">
                    <span style="font-size: 0.85rem; color: var(--text-secondary);">{acao}</span>
                    <span style="font-size: 0.85rem; font-weight: 600; color: var(--text-primary);">{count} ({pct:.1f}%)</span>
                </div>
                <div class="progress-container">
                    <div class="progress-bar {bar_class}" style="width: {pct}%;"></div>
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    with col_chart2:
        # Gráfico de distribuição por tempo até esgotar
        st.markdown("##### Por Tempo até Esgotar")
        
        tempo_col = ap.get("Tempo até esgotar (relatório)", ap.get("tempo_esgotar", pd.Series()))
        if not tempo_col.empty:
            tempo_counts = tempo_col.value_counts().head(8)
            
            for tempo, count in tempo_counts.items():
                pct = count / total_skus * 100
                tempo_str = str(tempo).strip()
                
                if "sem estoque" in tempo_str.lower():
                    bar_class = "critical"
                elif "1 semana" in tempo_str.lower() or "2 semana" in tempo_str.lower():
                    bar_class = "warning"
                else:
                    bar_class = "good"
                
                st.markdown(f"""
                <div style="margin-bottom: 0.75rem;">
                    <div style="display: flex; justify-content: space-between; margin-bottom: 0.25rem;">
                        <span style="font-size: 0.85rem; color: var(--text-secondary);">{tempo_str or 'N/A'}</span>
                        <span style="font-size: 0.85rem; font-weight: 600; color: var(--text-primary);">{count} ({pct:.1f}%)</span>
                    </div>
                    <div class="progress-container">
                        <div class="progress-bar {bar_class}" style="width: {pct}%;"></div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("Dados de tempo até esgotar não disponíveis.")
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # ===== CURVA ABC =====
    st.markdown("#### Curva ABC de Produtos")
    
    # Calcular curva ABC baseada em vendas
    abc_df = ap[ap["__vendas_30d"] > 0].copy()
    if not abc_df.empty:
        abc_df = abc_df.sort_values("__vendas_30d", ascending=False).reset_index(drop=True)
        abc_df["__vendas_acum"] = abc_df["__vendas_30d"].cumsum()
        total_vendas_abc = abc_df["__vendas_30d"].sum()
        abc_df["__pct_acum"] = abc_df["__vendas_acum"] / total_vendas_abc * 100
        
        # Classificar
        abc_df["Curva"] = np.where(
            abc_df["__pct_acum"] <= 80, "A",
            np.where(abc_df["__pct_acum"] <= 95, "B", "C")
        )
        
        curva_a = int((abc_df["Curva"] == "A").sum())
        curva_b = int((abc_df["Curva"] == "B").sum())
        curva_c = int((abc_df["Curva"] == "C").sum())
        
        vendas_a = abc_df.loc[abc_df["Curva"] == "A", "__vendas_30d"].sum()
        vendas_b = abc_df.loc[abc_df["Curva"] == "B", "__vendas_30d"].sum()
        vendas_c = abc_df.loc[abc_df["Curva"] == "C", "__vendas_30d"].sum()
        
        abc1, abc2, abc3 = st.columns(3)
        
        with abc1:
            st.markdown(f"""
            <div class="health-card good">
                <div style="display: flex; align-items: center; gap: 0.75rem;">
                    <span class="abc-badge a">A</span>
                    <div>
                        <div class="health-label">Curva A (80% vendas)</div>
                        <div class="health-value">{curva_a}</div>
                        <div class="health-subtitle">{vendas_a:.0f} un. vendidas</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with abc2:
            st.markdown(f"""
            <div class="health-card warning">
                <div style="display: flex; align-items: center; gap: 0.75rem;">
                    <span class="abc-badge b">B</span>
                    <div>
                        <div class="health-label">Curva B (15% vendas)</div>
                        <div class="health-value">{curva_b}</div>
                        <div class="health-subtitle">{vendas_b:.0f} un. vendidas</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with abc3:
            st.markdown(f"""
            <div class="health-card info">
                <div style="display: flex; align-items: center; gap: 0.75rem;">
                    <span class="abc-badge c">C</span>
                    <div>
                        <div class="health-label">Curva C (5% vendas)</div>
                        <div class="health-value">{curva_c}</div>
                        <div class="health-subtitle">{vendas_c:.0f} un. vendidas</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
        
        # Top 10 produtos
        st.markdown("#### Top 10 Produtos por Vendas")
        
        top10 = abc_df.head(10)[["SKU", "Produto", "__vendas_30d", "__estoque", "__dias_estoque", "Curva"]].copy()
        top10.columns = ["SKU", "Produto", "Vendas 30d", "Estoque", "Dias Estoque", "Curva"]
        top10["Dias Estoque"] = top10["Dias Estoque"].apply(lambda x: f"{x:.0f}" if x < 999 else "∞")
        
        for i, row in top10.iterrows():
            rank = i + 1
            dias_class = "critical" if row["Estoque"] == 0 else "warning" if float(row["Dias Estoque"].replace("∞", "999")) < 7 else "good"
            
            st.markdown(f"""
            <div class="priority-item">
                <div class="priority-rank">{rank}</div>
                <div class="priority-info">
                    <div class="priority-sku">{row['SKU']}</div>
                    <div class="priority-detail">{str(row['Produto'])[:50]}...</div>
                </div>
                <div style="text-align: center; padding: 0 1rem;">
                    <div style="font-size: 1.1rem; font-weight: 600; color: var(--accent);">{int(row['Vendas 30d'])}</div>
                    <div style="font-size: 0.7rem; color: var(--text-muted);">vendas/mês</div>
                </div>
                <div style="text-align: center; padding: 0 1rem;">
                    <div style="font-size: 1.1rem; font-weight: 600; color: var(--text-primary);">{int(row['Estoque'])}</div>
                    <div style="font-size: 0.7rem; color: var(--text-muted);">em estoque</div>
                </div>
                <div class="priority-action">
                    <div class="priority-qty" style="color: var(--primary-light);">{row['Curva']}</div>
                    <div class="priority-label">curva</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

# ============================================================
# TAB 2: Oportunidades
# ============================================================

with t2:
    st.markdown("### Oportunidades de Crescimento")
    
    if opps_full_report.empty and sales_opps.empty:
        st.info("Nenhuma oportunidade identificada no momento.")
    
    if not opps_full_report.empty:
        st.markdown("#### 🚀 Produtos com Giro mas sem Estoque no Full")
        st.markdown("""
        <div class="info-box">
            <p>Estes produtos tiveram vendas nos últimos 30 dias, mas estão com estoque zerado no Full. Priorize o envio destes itens para evitar perda de vendas.</p>
        </div>
        """, unsafe_allow_html=True)
        show_df(opps_full_report, height=400)
        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
        
    if not sales_opps.empty:
        st.markdown("#### 🎯 Oportunidades Fora do Full")
        st.markdown("""
        <div class="info-box">
            <p>Estes produtos estão vendendo bem fora do Full. Considere enviá-los para o Full para aumentar sua conversão e aproveitar os benefícios do programa.</p>
        </div>
        """, unsafe_allow_html=True)
        show_df(sales_opps, height=400)

# ============================================================
# TAB 3: Movimentações
# ============================================================

if t3 is not None:
    with t3:
        st.markdown("### Auditoria de Movimentações")
        
        if mov_summary:
            m1, m2, m3 = st.columns(3)
            m1.metric("Total Retiradas", _format_int(mov_summary["retiradas_abs"]))
            m2.metric("Total Vendidas", _format_int(mov_summary["vendidas_abs"]))
            
            diff = mov_summary["retiradas_abs"] - mov_summary["vendidas_abs"]
            m3.metric("Diferença", _format_int(diff), delta=_format_int(diff), delta_color="inverse" if diff > 0 else "normal")
            
            st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
            
            if not audit.empty:
                st.markdown("#### Divergências Identificadas")
                divergencias = audit[~audit["_ok"]].copy()
                if not divergencias.empty:
                    st.warning(f"Foram encontradas {len(divergencias)} divergências significativas entre retiradas e vendas.")
                    show_df(divergencias, height=400)
                else:
                    st.success("✅ Nenhuma divergência significativa encontrada nas movimentações.")
                
                st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
                st.markdown("#### Histórico Completo")
                show_df(mov, height=500)
        else:
            st.info("Carregue o consolidado de movimentações para ver a auditoria.")

# ============================================================
# TAB 4: Painel de Ação
# ============================================================

with t4:
    st.markdown("### Painel de Ação e Reposição")
    
    st.markdown("""
    <div class="info-box">
        <p>Lista completa de SKUs com sugestões de envio baseadas na cobertura desejada e crescimento esperado.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Filtros rápidos
    f1, f2 = st.columns(2)
    with f1:
        filtro_acao = st.multiselect(
            "Filtrar por Ação",
            options=list(action_panel["Ação Recomendada"].unique()),
            default=list(action_panel["Ação Recomendada"].unique())
        )
    
    with f2:
        search_sku = st.text_input("Buscar SKU ou Produto", "")
    
    # Aplicar filtros
    display_panel = action_panel[action_panel["Ação Recomendada"].isin(filtro_acao)].copy()
    if search_sku:
        display_panel = display_panel[
            display_panel["SKU"].str.contains(search_sku, case=False, na=False) |
            display_panel["Produto"].astype(str).str.contains(search_sku, case=False, na=False)
        ]
    
    # Ordenar
    display_panel = display_panel.sort_values(["prioridade", "Sugestão Envio (un.)"], ascending=[True, False])
    
    # Formatação para exibição
    cols_to_show = ["SKU", "Produto", "Estoque Atual", "Vendas Período", "A caminho", "Sugestão Envio (un.)", "Ação Recomendada"]
    if "days_since_last_sale" in display_panel.columns:
        cols_to_show.append("days_since_last_sale")
        
    show_df(display_panel[cols_to_show], height=600)

# ============================================================
# TAB 7: Planejador de Envios
# ============================================================

with t7:
    st.markdown("### Planejador de Envios")
    
    st.markdown("""
    <div class="info-box">
        <p>Use esta ferramenta para planejar seus próximos envios ao Full com base em dados reais de vendas e estoque.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Preparar dados
    plan_df = action_panel.copy()
    plan_df["__estoque"] = pd.to_numeric(plan_df.get("Estoque Atual", 0), errors="coerce").fillna(0.0)
    plan_df["__vendas_30d"] = pd.to_numeric(plan_df.get("Vendas 30d", plan_df.get("Vendas Período", 0)), errors="coerce").fillna(0.0)
    plan_df["__vel"] = pd.to_numeric(plan_df.get("Velocidade Média Dia", 0), errors="coerce").fillna(0.0)
    plan_df["__a_caminho"] = pd.to_numeric(plan_df.get("A caminho", 0), errors="coerce").fillna(0.0)
    plan_df["__sug_original"] = pd.to_numeric(plan_df.get("Sugestão Envio (un.)", 0), errors="coerce").fillna(0.0)
    
    # Calcular velocidade se não existir
    if (plan_df["__vel"] == 0).all():
        plan_df["__vel"] = plan_df["__vendas_30d"] / 30
    
    # ===== CALCULADORA DE REPOSIÇÃO =====
    st.markdown("#### Calculadora de Reposição")
    
    calc_col1, calc_col2, calc_col3 = st.columns(3)
    
    with calc_col1:
        cobertura_desejada = st.number_input(
            "Cobertura desejada (dias)",
            min_value=7,
            max_value=90,
            value=int(cover_days_effective),
            step=1,
            help="Quantos dias de estoque você deseja ter após o envio"
        )
    
    with calc_col2:
        margem_seguranca = st.slider(
            "Margem de segurança (%)",
            min_value=0,
            max_value=50,
            value=10,
            step=5,
            help="Adiciona uma margem extra à sugestão para imprevistos"
        )
    
    with calc_col3:
        filtro_prioridade = st.multiselect(
            "Filtrar por prioridade",
            options=["Apenas rupturas", "Rupturas + Críticos", "Todos com sugestão"],
            default=["Todos com sugestão"]
        )
    
    # Calcular sugestões
    plan_df["__dias_estoque"] = np.where(
        plan_df["__vel"] > 0,
        (plan_df["__estoque"] + plan_df["__a_caminho"]) / plan_df["__vel"],
        np.where(plan_df["__estoque"] > 0, 999, 0)
    )
    
    # Sugestão = (cobertura * velocidade * (1 + margem)) - estoque - a_caminho
    margem_mult = 1 + (margem_seguranca / 100)
    plan_df["__sug_calc"] = np.ceil(
        (cobertura_desejada * plan_df["__vel"] * margem_mult) - plan_df["__estoque"] - plan_df["__a_caminho"]
    )
    plan_df["__sug_calc"] = np.maximum(0, plan_df["__sug_calc"])
    
    # Aplicar filtros
    mask = plan_df["__sug_calc"] > 0
    
    if "Apenas rupturas" in filtro_prioridade:
        mask = mask & (plan_df["__estoque"] <= 0) & (plan_df["__vendas_30d"] > 0)
    elif "Rupturas + Críticos" in filtro_prioridade:
        mask = mask & ((plan_df["__estoque"] <= 0) | (plan_df["__dias_estoque"] < 7))
    
    envio_df = plan_df[mask].copy()
    
    # Ordenar por prioridade (rupturas primeiro, depois por vendas)
    envio_df["__prio_sort"] = np.where(
        envio_df["__estoque"] <= 0, 0,
        np.where(envio_df["__dias_estoque"] < 7, 1, 2)
    )
    envio_df = envio_df.sort_values(["__prio_sort", "__vendas_30d"], ascending=[True, False]).reset_index(drop=True)
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # ===== RESUMO DO ENVIO =====
    st.markdown("#### Resumo do Envio Sugerido")
    
    total_skus_envio = len(envio_df)
    total_unidades = int(envio_df["__sug_calc"].sum())
    rupturas_resolvidas = int(((envio_df["__estoque"] <= 0) & (envio_df["__vendas_30d"] > 0)).sum())
    
    r1, r2, r3, r4 = st.columns(4)
    
    with r1:
        st.markdown(f"""
        <div class="health-card info">
            <div class="health-label">SKUs para Enviar</div>
            <div class="health-value">{total_skus_envio}</div>
            <div class="health-subtitle">produtos diferentes</div>
        </div>
        """, unsafe_allow_html=True)
    
    with r2:
        st.markdown(f"""
        <div class="health-card info">
            <div class="health-label">Total de Unidades</div>
            <div class="health-value">{total_unidades:,}</div>
            <div class="health-subtitle">unidades sugeridas</div>
        </div>
        """, unsafe_allow_html=True)
    
    with r3:
        st.markdown(f"""
        <div class="health-card good">
            <div class="health-label">Rupturas Resolvidas</div>
            <div class="health-value">{rupturas_resolvidas}</div>
            <div class="health-subtitle">SKUs voltarão a ter estoque</div>
        </div>
        """, unsafe_allow_html=True)
    
    with r4:
        st.markdown(f"""
        <div class="health-card info">
            <div class="health-label">Cobertura Pós-Envio</div>
            <div class="health-value">{cobertura_desejada}</div>
            <div class="health-subtitle">dias de estoque</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # ===== LISTA PRIORIZADA =====
    st.markdown("#### Lista Priorizada de Envio")
    
    if envio_df.empty:
        st.success("✅ Nenhum produto precisa de reposição com os parâmetros atuais!")
    else:
        # Mostrar top 20 como cards
        st.markdown("##### Top 20 Prioridades")
        
        for i, row in envio_df.head(20).iterrows():
            rank = envio_df.index.get_loc(i) + 1
            sku = row.get("SKU", "N/A")
            produto = str(row.get("Produto", ""))[:40]
            estoque = int(row["__estoque"])
            vendas = int(row["__vendas_30d"])
            sugestao = int(row["__sug_calc"])
            dias = row["__dias_estoque"]
            
            if estoque <= 0:
                status_class = "critical"
                status_icon = "🚨"
                status_text = "RUPTURA"
            elif dias < 7:
                status_class = "warning"
                status_icon = "⚠️"
                status_text = f"{dias:.0f} dias"
            else:
                status_class = "info"
                status_icon = "📦"
                status_text = f"{dias:.0f} dias"
            
            st.markdown(f"""
            <div class="priority-item">
                <div class="priority-rank">{rank}</div>
                <div class="priority-info">
                    <div class="priority-sku">{sku}</div>
                    <div class="priority-detail">{produto}...</div>
                </div>
                <div style="text-align: center; padding: 0 0.75rem;">
                    <div style="font-size: 0.9rem; font-weight: 600; color: var(--text-primary);">{estoque}</div>
                    <div style="font-size: 0.65rem; color: var(--text-muted);">estoque</div>
                </div>
                <div style="text-align: center; padding: 0 0.75rem;">
                    <div style="font-size: 0.9rem; font-weight: 600; color: var(--accent);">{vendas}</div>
                    <div style="font-size: 0.65rem; color: var(--text-muted);">vendas/mês</div>
                </div>
                <div style="text-align: center; padding: 0 0.75rem;">
                    <span class="alert-badge {status_class}">{status_icon} {status_text}</span>
                </div>
                <div class="priority-action">
                    <div class="priority-qty">{sugestao}</div>
                    <div class="priority-label">enviar</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
        
        # Tabela completa
        st.markdown("##### Lista Completa")
        
        lista_envio = envio_df[["SKU", "Produto", "__estoque", "__vendas_30d", "__a_caminho", "__dias_estoque", "__sug_calc"]].copy()
        lista_envio.columns = ["SKU", "Produto", "Estoque Atual", "Vendas 30d", "A Caminho", "Dias Estoque", "Sugestão Envio"]
        lista_envio["Dias Estoque"] = lista_envio["Dias Estoque"].apply(lambda x: f"{x:.0f}" if x < 999 else "∞")
        lista_envio["Estoque Atual"] = lista_envio["Estoque Atual"].astype(int)
        lista_envio["Vendas 30d"] = lista_envio["Vendas 30d"].astype(int)
        lista_envio["A Caminho"] = lista_envio["A Caminho"].astype(int)
        lista_envio["Sugestão Envio"] = lista_envio["Sugestão Envio"].astype(int)
        
        show_df_minimal(lista_envio, height=400)
        
        # Botão de download
        st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
        
        col_dl1, col_dl2 = st.columns(2)
        
        with col_dl1:
            xlsx_envio = to_xlsx_bytes({"Lista de Envio": safe_df(lista_envio)})
            st.download_button(
                "📥 Baixar Lista de Envio (Excel)",
                data=xlsx_envio,
                file_name=f"lista_envio_full_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        
        with col_dl2:
            # Criar resumo simples para envio
            resumo_envio = lista_envio[["SKU", "Sugestão Envio"]].copy()
            xlsx_resumo = to_xlsx_bytes({"Resumo de Envio": safe_df(resumo_envio)})
            st.download_button(
                "📥 Baixar Resumo (Excel)",
                data=xlsx_resumo,
                file_name=f"resumo_envio_full_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    
    st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
    
    # ===== SIMULADOR DE IMPACTO =====
    st.markdown("#### Simulador de Impacto")
    
    st.markdown("""
    <div class="info-box">
        <p>Simule o impacto do envio na saúde do seu estoque.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Métricas atuais vs pós-envio
    total_skus_sim = len(plan_df)
    
    # Atual
    rupturas_atual = int(((plan_df["__estoque"] <= 0) & (plan_df["__vendas_30d"] > 0)).sum())
    taxa_ruptura_atual = (rupturas_atual / total_skus_sim * 100) if total_skus_sim > 0 else 0
    
    # Pós-envio (simulado)
    plan_df["__estoque_pos"] = plan_df["__estoque"] + plan_df["__sug_calc"]
    rupturas_pos = int(((plan_df["__estoque_pos"] <= 0) & (plan_df["__vendas_30d"] > 0)).sum())
    taxa_ruptura_pos = (rupturas_pos / total_skus_sim * 100) if total_skus_sim > 0 else 0
    
    sim1, sim2 = st.columns(2)
    
    with sim1:
        st.markdown("##### Situação Atual")
        st.markdown(f"""
        <div class="health-card critical">
            <div class="health-label">Taxa de Ruptura</div>
            <div class="health-value">{taxa_ruptura_atual:.1f}%</div>
            <div class="health-subtitle">{rupturas_atual} SKUs em ruptura</div>
        </div>
        """, unsafe_allow_html=True)
    
    with sim2:
        st.markdown("##### Pós-Envio Sugerido")
        pos_class = "good" if taxa_ruptura_pos < 5 else "warning" if taxa_ruptura_pos < 15 else "critical"
        st.markdown(f"""
        <div class="health-card {pos_class}">
            <div class="health-label">Taxa de Ruptura</div>
            <div class="health-value">{taxa_ruptura_pos:.1f}%</div>
            <div class="health-subtitle">{rupturas_pos} SKUs em ruptura</div>
        </div>
        """, unsafe_allow_html=True)
    
    # Melhoria
    melhoria = taxa_ruptura_atual - taxa_ruptura_pos
    if melhoria > 0:
        st.success(f"✅ O envio sugerido reduzirá a taxa de ruptura em **{melhoria:.1f} pontos percentuais**!")
    elif melhoria == 0:
        st.info("ℹ️ O envio manterá a taxa de ruptura atual.")
    else:
        st.warning("⚠️ Verifique os parâmetros - a taxa de ruptura não melhorará.")


# ============================================================
# TAB 5: Exportar
# ============================================================

with t5:
    st.markdown("### Exportar Dados")
    
    st.markdown("""
    <div class="info-box">
        <p>Exporte os dados do painel em diferentes formatos para uso externo.</p>
    </div>
    """, unsafe_allow_html=True)

    col_exp1, col_exp2 = st.columns(2)
    
    with col_exp1:
        st.markdown("#### Excel do Painel")
        xlsx_painel = to_xlsx_bytes({"Painel de Ação": safe_df(action_panel)})
        st.download_button(
            "📥 Baixar Excel do Painel",
            data=xlsx_painel,
            file_name="painel_de_acao_full.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col_exp2:
        st.markdown("#### Excel Completo")
        
        def _build_config_df() -> pd.DataFrame:
            sales_loaded = bool(sales_raw is not None and sales_map is not None)

            period_txt = ""
            if mov_summary is not None and mov_summary.get("period_start") and mov_summary.get("period_end"):
                period_txt = f"{mov_summary['period_start'].strftime('%d/%m/%Y')} a {mov_summary['period_end'].strftime('%d/%m/%Y')}"
            else:
                period_txt = "Modo leve (sem consolidado)"

            rows = [
                ("Gerado em", datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y %H:%M:%S")),
                ("Data base", today_sp.strftime("%d/%m/%Y")),
                ("Modo", "Leve" if mov is None else "Completo"),
                ("Modo sazonal", season.get("mode", "")),
                ("Contexto sazonal", season.get("context", "")),
                ("Usou calibração sazonal", "Sim" if use_seasonal else "Não"),
                ("Dias de cobertura aplicados", str(int(cover_days_effective))),
                ("Crescimento esperado (%)", str(round(float(growth_pct) * 100.0, 2))),
                ("Mínimo vendas (fora do Full)", str(float(min_sales_qty))),
                ("Top N (fora do Full)", str(int(top_n))),
                ("Excluir canceladas e devoluções", "Sim" if bool(exclude_canceled_and_returns) else "Não"),
                ("Relatório de vendas carregado", "Sim" if sales_loaded else "Não"),
                ("Aba do relatório de vendas", (sales_map.get("sheet", "") if sales_loaded else "")),
                ("Cabeçalho detectado automaticamente", (str(bool(sales_map.get("header_auto", False))) if sales_loaded else "")),
                ("Período do consolidado", period_txt),
            ]
            return pd.DataFrame(rows, columns=["Parâmetro", "Valor"])

        if mov is None:
            audit_export = pd.DataFrame([{"info": "Modo leve: sem consolidado de movimentações, auditoria indisponível."}])
        else:
            audit_export = audit if audit is not None and not audit.empty else pd.DataFrame([{"info": "Sem divergências relevantes."}])

        has_opps = sales_opps is not None and not sales_opps.empty
        
        panel_sorted = action_panel.copy()
        if "prioridade" in panel_sorted.columns:
            if "Sugestão Envio (un.)" in panel_sorted.columns:
                panel_sorted = panel_sorted.sort_values(by=["prioridade", "Sugestão Envio (un.)"], ascending=[True, False])
            else:
                panel_sorted = panel_sorted.sort_values(by=["prioridade"], ascending=[True])

        xlsx_completo = to_xlsx_bytes({
            "Painel de ação": safe_df(panel_sorted),
            "Auditoria": safe_df(audit_export),
            "Config": safe_df(_build_config_df()),
            "Oportunidades fora do Full": safe_df(sales_opps) if has_opps else pd.DataFrame()
        })

        st.download_button(
            "📥 Baixar Excel Completo",
            data=xlsx_completo,
            file_name="arquivo_final_full.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# Footer
st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
st.markdown(f"""
<div style="text-align: center; padding: 2rem 0;">
    <p style="color: var(--text-primary); font-size: 1rem; font-weight: 600;">
        © Desenvolvido por Vinicius Lima / CNPJ: 47.192.694/0001-70
    </p>
</div>
""", unsafe_allow_html=True)

# ============================================================
# TAB 8: Guia de Uso
# ============================================================

with t8:
    show_guia_uso()
